import os
import requests
import asyncio
import sqlite3
import telebot
from telebot.async_telebot import AsyncTeleBot
from telebot import types
import traceback
import datetime
import aiohttp
import random
import time

from elschool_client import ElschoolClient

bot = AsyncTeleBot('')

VERSION = "v.3.0.1-Beta"
SP_DAYS = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
USERS_IN_CYCLE = 0
AVG_PER_USER = 0.0
IMAGES = {}

connect = sqlite3.connect('els3_test.db', check_same_thread=False)
cursor = connect.cursor()

cursor.execute("""CREATE TABLE IF NOT EXISTS all_users(
    user_id INTEGER,
    name TEXT,
    username TEXT
    )
""")
connect.commit()

cursor.execute("""CREATE TABLE IF NOT EXISTS credits(
    user_id INTEGER,
    jwtoken TEXT,
    refreshtoken TEXT
    )
""")
connect.commit()

cursor.execute("""CREATE TABLE IF NOT EXISTS states(
    user_id INTEGER,
    type_of_state TEXT
    )
""")
connect.commit()

cursor.execute("""CREATE TABLE IF NOT EXISTS time(
    user_id INTEGER,
    time_marks INTEGER,
    time_tabel INTEGER
    )
""")
connect.commit()

cursor.execute("""CREATE TABLE IF NOT EXISTS marks(
    mark_id TEXT,
    user_id INTEGER,
    mark_value INTEGER,
    predmet TEXT,
    datetime_mark TEXT
    )
""")
connect.commit()

cursor.execute("""CREATE TABLE IF NOT EXISTS users_posting(
    user_id INTEGER
    )
""")
connect.commit()


@bot.message_handler(commands=['start'])
async def start_func(message):
    is_new = cursor.execute("SELECT * FROM all_users WHERE user_id=?", (message.from_user.id,)).fetchone()
    if is_new is None:
        user = message.from_user
        name1, name2 = user.first_name.replace('[', '').replace(']', ''), user.last_name
        if name2 is None:
            name2 = ''
        else:
            name2 = ' ' + name2.replace('[', '').replace(']', '')
        cursor.execute("INSERT INTO all_users VALUES(?,?,?);", [message.from_user.id, f'{name1}{name2}', user.username])
        connect.commit()
    else:
        try:
            await bot.delete_message(message.from_user.id, message.message_id)
        except Exception:
            pass

    menu = types.InlineKeyboardMarkup()
    menu.add(types.InlineKeyboardButton(text='👤 Личный кабинет', callback_data='profile'))
    menu.add(types.InlineKeyboardButton(text='🎖 Оценки', callback_data='mm'))
    menu.add(types.InlineKeyboardButton(text='📚 ДЗ', callback_data='dz'))
    menu.add(types.InlineKeyboardButton(text='🤖 Состояние автопарсинга', callback_data='status_parsing'))
    menu.add(types.InlineKeyboardButton(text='🛠 Техподдержка', callback_data='help'), types.InlineKeyboardButton(text='👨‍💻Исходный код', url='https://github.com/theslothbear/Elschool-Help-Bot'))

    await bot.send_photo(message.from_user.id, photo='https://imgur.com/nbDpsEi.jpg', caption=f'🏠*Главное меню Elschool Help Bot ({VERSION})*', parse_mode='markdown', reply_markup=menu)


@bot.callback_query_handler(lambda call: call.data == 'menu')
async def menu_func(call):
    menu = types.InlineKeyboardMarkup()
    menu.add(types.InlineKeyboardButton(text='👤 Личный кабинет', callback_data='profile'))
    menu.add(types.InlineKeyboardButton(text='🎖 Оценки', callback_data='mm'))
    menu.add(types.InlineKeyboardButton(text='📚 ДЗ', callback_data='dz'))
    menu.add(types.InlineKeyboardButton(text='🤖 Состояние автопарсинга', callback_data='status_parsing'))
    menu.add(types.InlineKeyboardButton(text='🛠 Техподдержка', callback_data='help'), types.InlineKeyboardButton(text='👨‍💻Исходный код', url='https://github.com/theslothbear/Elschool-Help-Bot'))

    try:
        await bot.edit_message_media(chat_id=call.from_user.id, message_id=call.message.message_id, media=types.InputMediaPhoto(media=IMAGES['logo'], caption=f'🏠*Главное меню Elschool Help Bot ({VERSION})*', parse_mode='markdown'), reply_markup=menu)
    except Exception:
        await bot.send_photo(call.from_user.id, photo='https://imgur.com/nbDpsEi.jpg', caption=f'🏠*Главное меню Elschool Help Bot ({VERSION})*', parse_mode='markdown', reply_markup=menu)
        try:
            await bot.delete_message(call.from_user.id, call.message.message_id)
        except Exception:
            pass


@bot.message_handler(commands=['privacy'])
async def privacy_func(message):
    priv = types.InlineKeyboardMarkup()
    priv.add(types.InlineKeyboardButton(text='📗Пользовательское соглашение', url='https://telegra.ph/Polzovatelskoe-soglashenie-Elschool-Help-Bot-09-04'))
    priv.add(types.InlineKeyboardButton(text='🔙В меню', callback_data='menu'))
    await bot.send_message(message.from_user.id, '*Пункт 2.2 Пользовательского соглашения:*\n\n_Начиная использовать Бот/его отдельные функции, Пользователь считается принявшим условия Соглашения в полном объеме без всяких оговорок и исключений._', reply_markup=priv, parse_mode='markdown')

@bot.callback_query_handler(lambda call: call.data == 'profile')
async def profile_func(call):
    credits = cursor.execute("SELECT * FROM credits WHERE user_id=?", (call.from_user.id,)).fetchone()
    pr = types.InlineKeyboardMarkup()
    if credits is None:
        pr.add(types.InlineKeyboardButton(text='➕ Подключить аккаунт ELSCHOOL', callback_data='podkl'))
        t1, t2 = 'Не задано', 'Не задано'
    else:
        pr.add(types.InlineKeyboardButton(text='✏ Изменить аккаунт ELSCHOOL', callback_data='podkl'))
        # pr.add(types.InlineKeyboardButton(text='💦Дополнительные функции', callback_data='dop'))
        # pr.add(types.InlineKeyboardButton(text='🤖 Состояние автопарсинга', callback_data='status_parsing'))
        pr.add(types.InlineKeyboardButton(text='🗑 Удалить свои данные', callback_data='DATA_DELETE'))
        t1, t2 = credits[1], credits[2]

    n1 = call.from_user.first_name
    if n1 is None:
        n1 = ''
    n2 = call.from_user.last_name
    if n2 is None:
        n2 = ''

    pr.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='menu'))

    try:
        await bot.edit_message_media(chat_id=call.from_user.id, message_id=call.message.message_id, media=types.InputMediaPhoto(media=IMAGES['user'], caption=f'<b>👤Профиль {n1} {n2}</b>\n\n💠JWToken: <span class="tg-spoiler">{t1}</span>\n\n🔐RefreshToken: <span class="tg-spoiler">{t2}</span>', parse_mode='HTML'), reply_markup=pr)
    except Exception:
        await bot.send_photo(call.from_user.id, photo='https://imgur.com/ocHQUkF.jpg', caption=f'<b>👤Профиль {n1} {n2}</b>\n\n💠JWToken: <span class="tg-spoiler">{t1}</span>\n\n🔐RefreshToken: <span class="tg-spoiler">{t2}</span>', parse_mode='HTML', reply_markup=pr)
        try:
            await bot.delete_message(call.from_user.id, call.message.message_id)
        except Exception:
            pass


@bot.callback_query_handler(lambda call: call.data == 'status_parsing')
async def status_parsing(call):
    r = cursor.execute("SELECT * FROM users_posting WHERE user_id=?", (call.from_user.id,)).fetchone()
    rm = types.InlineKeyboardMarkup()
    if r is None:
        rm.add(types.InlineKeyboardButton(text='✅ Включить автопроверку оценок', callback_data='enter_posting'))
        s = 'На данный момент вы не находитесь в пуле автоматического парсинга'
    else:
        rm.add(types.InlineKeyboardButton(text='❎ Выключить автопроверку оценок', callback_data='off_posting'))
        s = 'Вы находитесь в пуле автоматического парсинга'
    rm.add(types.InlineKeyboardButton(text='🔙Назад', callback_data='menu'))

    await bot.send_message(call.from_user.id, f'*🤖 Состояние автопарсинга*\n\n• Количество пользователей в пуле: {USERS_IN_CYCLE}\n• Среднее время на пользователя: {AVG_PER_USER} секунд\n\n_{s}_', reply_markup=rm, parse_mode='markdown')
    try:
        await bot.delete_message(call.from_user.id, call.message.message_id)
    except Exception:
        pass


@bot.callback_query_handler(lambda call: call.data == 'enter_posting')
async def enter_posting_func(call):
    r = cursor.execute("SELECT * FROM users_posting WHERE user_id=?", (call.from_user.id,)).fetchone()
    if r is None:
        global USERS_IN_CYCLE
        cursor.execute("INSERT INTO users_posting VALUES(?);", [call.from_user.id])
        connect.commit()

        USERS_IN_CYCLE += 1

    await status_parsing(call)


@bot.callback_query_handler(lambda call: call.data == 'off_posting')
async def off_posting_func(call):
    r = cursor.execute("SELECT * FROM users_posting WHERE user_id=?", (call.from_user.id,)).fetchone()
    if r is not None:
        global USERS_IN_CYCLE
        cursor.execute("DELETE FROM users_posting WHERE user_id=?", (call.from_user.id,))
        connect.commit()

        USERS_IN_CYCLE -= 1

    await status_parsing(call)


@bot.callback_query_handler(lambda call: call.data == 'DATA_DELETE')
async def delete_data_func(call):
    cursor.execute("DELETE FROM credits WHERE user_id=?", (call.from_user.id,))
    connect.commit()

    delmenu = types.InlineKeyboardMarkup()
    delmenu.add(types.InlineKeyboardButton(text='🔙 В меню', callback_data='menu'))

    await bot.send_message(call.from_user.id, '*✅ Ваши данные удалены*', parse_mode='markdown', reply_markup=delmenu)


@bot.callback_query_handler(lambda call: call.data == 'delete')
async def delete(call):
    try:
        await bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception:
        pass


@bot.callback_query_handler(lambda call: call.data == 'podkl')
async def podkl_func(call):
    ty = types.InlineKeyboardMarkup()
    ty.add(types.InlineKeyboardButton(text='📗 Пользовательское соглашение', url='https://telegra.ph/Polzovatelskoe-soglashenie-Elschool-Help-Bot-09-04'))
    ty.add(types.InlineKeyboardButton(text='➡ Продолжить', callback_data='podklok'))
    ty.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='menu'))

    await bot.send_message(call.from_user.id, f'🌐Соединение с аккаунтом *ELSCHOOL*\n\n❗Нажимая кнопку "Продолжить", Вы *подтверждаете*, что ознакомились с пользовательским соглашением.', reply_markup=ty, parse_mode='markdown')


@bot.callback_query_handler(lambda call: call.data == 'podklok')
async def podklok_func(call):
    cursor.execute("DELETE FROM states WHERE user_id=?", (call.from_user.id,))
    connect.commit()
    sdf = [call.from_user.id, 'login-and-password']
    cursor.execute("INSERT INTO states VALUES(?,?);", sdf)
    connect.commit()

    gde = types.InlineKeyboardMarkup()
    gde.add(types.InlineKeyboardButton(text='👉 Где взять токены', url='https://telegra.ph/Poluchenie-cookies-dlya-ElsHelp-09-04'))

    await bot.send_message(call.from_user.id, '🌐Соединение с аккаунтом *ELSCHOOL*\n\nВведите JWToken и RefreshToken, каждое в новой строке\n\n_Пример:_\n_aaaaa_\n_bbbbb_\n\n*Важно:* привязывать необходимо аккаунт *ученика!*', parse_mode='markdown', reply_markup=gde)


async def create_new_id():
    cursor.execute("SELECT * FROM marks")
    records = cursor.fetchall()
    new_id = len(records) + 1
    return new_id


@bot.message_handler(commands=['get_marks'])
async def get_marks_func(message):
    await process_marks(message.from_user.id, True)


async def process_marks(user_id, by_user=False):
    credits = cursor.execute("SELECT * FROM credits WHERE user_id=?", (user_id,)).fetchone()
    if credits is None:
        await bot.send_message(user_id, '❎ Вы не авторизованы')
        return False  # удалить из очереди парсинга

    if by_user:
        last_time = cursor.execute("SELECT * FROM time WHERE user_id=?", (user_id,)).fetchone()
        now_time = time.time()
        if (last_time is not None) and (now_time - last_time[1] < 300):
            await bot.reply_to(user_id, f'❎ Вы уже использовали эту функцию в последние 5 минут. Попробуйте через {int(5-(now_time - last_time[1])//60)} минут')
            return

        if last_time is None:
            last_time = [0, 0, 0]

        await bot.send_chat_action(user_id, 'typing')

    # Заранее извиняюсь за дальнейший код. Но переписывать не собираюсь. Себе дороже
    am = 0
    rowrt = [user_id]

    try:
        F, flag_norm = True, True
        cursor.execute(f"""CREATE TABLE IF NOT EXISTS t{rowrt[0]}(
            predmet TEXT,
            m5_marks TEXT,
            m4_marks TEXT,
            m3_marks TEXT,
            m2_marks TEXT,
            m1_marks TEXT,
            str_marks TEXT
            )
            """)
        connect.commit()

        r1239 = [0, 'RU']

        async with ElschoolClient(credits[1], credits[2]) as client:
            status0, t0 = await client.auth()
            if not status0:
                await bot.send_message(user_id, '❎ Не удалось войти в аккаунт Elschool. Возможно, стоит обновить cookies?')
                return False  # удалить из очереди парсинга

            status1, t1 = await client.get_url('https://elschool.ru/users/diaries')
            if not status1:
                await bot.send_message(user_id, '❎ Не удалось загрузить страницу дневника. Возможно, стоит обновить cookies?')
                return False  # удалить из очереди парсинга

            s = t1.split('class="btn">Табель</a>')[0].split(r'href="')[-1].split(r'"')[0]
            status2, t2 = await client.get_url(f'https://elschool.ru/users/diaries/{s}')
            if status2 is None:
                await bot.send_message(user_id, '❎ Не удалось загрузить страницу табеля. Возможно, стоит обновить cookies?')
                return False  # удалить из очереди парсинга

            r2 = t2

            spg, fl, col4 = [], True, -1
            for i in range(1, 100):
                str_marks = ''
                s1 = list(r2.split(f'<tbody period="{i}"'))
                if len(s1) > 1:
                    pr = s1[0].split(r'<th colspan="')[-1].split('>')[1].split('<')[0]
                    spo = []
                    l1 = list(s1[1].split(r'<td class="grades-period-name">1')[1].split('<span>'))
                    if s1[1].split(r'<td class="grades-period-name">1')[1][1:4] == 'чет':
                        col4 = s1[1].split(r'<td class="grades-period-name">4')[1].split('<td class="grades-period-name">1')[0].count('<span>')
                        col3 = s1[1].split(r'<td class="grades-period-name">3')[1].split('<td class="grades-period-name">1')[0].count('<span>') - col4
                        col2 = s1[1].split(r'<td class="grades-period-name">2')[1].split('<td class="grades-period-name">1')[0].count('<span>') - col3 - col4
                        col1 = s1[1].split(r'<td class="grades-period-name">1')[1].split('<td class="grades-period-name">1')[0].count('<span>') - col2 - col3 - col4
                        #print(f'{pr}: 1 четверть - {col1}, 2 четверть - {col2}, 3 четверть - {col3}, 4 четверть - {col4}')
                    elif s1[1].split(r'<td class="grades-period-name">1')[1][1:4] == 'три':
                        col3 = s1[1].split(r'<td class="grades-period-name">3')[1].split('<td class="grades-period-name">1')[0].count('<span>')
                        col2 = s1[1].split(r'<td class="grades-period-name">2')[1].split('<td class="grades-period-name">1')[0].count('<span>') - col3
                        col1 = s1[1].split(r'<td class="grades-period-name">1')[1].split('<td class="grades-period-name">1')[0].count('<span>') - col2 - col3
                    elif s1[1].split(r'<td class="grades-period-name">1')[1][1:4] == 'пол':
                        #await bot.send_message(-1001984000978, f'{str(await bot.get_chat_member(rowrt[0], rowrt[0]))}')
                        col4 = 0
                        col3 = s1[1].split(r'<td class="grades-period-name">2')[1].split('<td class="grades-period-name">1')[0].count('<span>') - col4
                        col2 = 0
                        col1 = s1[1].split(r'<td class="grades-period-name">1')[1].split('<td class="grades-period-name">1')[0].count('<span>') - col2 - col3 - col4
                    for r in l1[1:]:
                        yu = r.split('</span>')[0]
                        spo.append(yu)
                        str_marks += f'{yu} '
                    if col4 != -1 and flag_norm:
                        spg.append({'Предмет': f'{pr}', 'Оценки': f'{" ".join(spo)}', 'Colvo': f'{col1} {col2} {col3} {col4}', 'str': str_marks[0:len(str_marks)-1]})
                    elif flag_norm:
                        spg.append({'Предмет': f'{pr}', 'Оценки': f'{" ".join(spo)}', 'Colvo': f'{col1} {col2} {col3}', 'str': str_marks[0:len(str_marks)-1]})

                else:
                    if i == 1:
                        F = False
                    break
            if spg == [] and F and flag_norm:
                if r1239[1] == 'RU':
                    tyi = types.InlineKeyboardMarkup()
                    pr1 = types.InlineKeyboardButton(text = '✏Изменить аккаунт ELSCHOOL', callback_data='podkl')
                    zx2 = types.InlineKeyboardButton(text = '🔙В меню', callback_data = 'menu')
                    tyi.add(pr1)
                    tyi.add(zx2)
                    await bot.send_message(rowrt[0], '❌*Ошибка!* \nПохоже, логин либо пароль введены неправильно😿', reply_markup=tyi, parse_mode='markdown')
                elif r1239[1] == 'EN':
                    yi = types.InlineKeyboardMarkup()
                    pr1 = types.InlineKeyboardButton(text = '✏Change ELSCHOOL account', callback_data='podkl')
                    zx2 = types.InlineKeyboardButton(text = '🔙Menu', callback_data = 'menu')
                    tyi.add(pr1)
                    tyi.add(zx2)
                    await bot.send_message(rowrt[0], '❌*Error!* \n Looks like the username or password entered incorrectly😿\n\n', reply_markup=tyi, parse_mode='markdown')
            elif F and flag_norm:
                cursor.execute(f"SELECT * FROM t{rowrt[0]}")
                records = cursor.fetchall()
                cursor.execute(f"DELETE FROM t{rowrt[0]}")
                connect.commit()
                if col4 != -1:
                    for s in spg:
                        m1m_1, m2m_1, m3m_1, m4m_1, m5m_1, n = 0, 0, 0, 0, 0, 1
                        m1m_2, m2m_2, m3m_2, m4m_2, m5m_2 = 0, 0, 0, 0, 0
                        m1m_3, m2m_3, m3m_3, m4m_3, m5m_3 = 0, 0, 0, 0, 0
                        m1m_4, m2m_4, m3m_4, m4m_4, m5m_4 = 0, 0, 0, 0, 0
                        c1, c2, c3, c4 = map(int, s['Colvo'].split())
                        for mark in list(s['Оценки'].split()):
                            if int(mark) == 5:
                                if n <= c1:
                                    m5m_1 +=1
                                elif n <= c1 + c2:
                                    m5m_2 += 1
                                elif n <= c1 + c2 + c3:
                                    m5m_3 += 1
                                else:
                                    m5m_4 += 1
                                n+=1
                            elif int(mark) == 4:
                                if n <= c1:
                                    m4m_1 +=1
                                elif n <= c1 + c2:
                                    m4m_2 += 1
                                elif n <= c1 + c2 + c3:
                                    m4m_3 += 1
                                else:
                                    m4m_4 += 1
                                n+=1
                            elif int(mark) == 3:
                                if n <= c1:
                                    m3m_1 +=1
                                elif n <= c1 + c2:
                                    m3m_2 += 1
                                elif n <= c1 + c2 + c3:
                                    m3m_3 += 1
                                else:
                                    m3m_4 += 1
                                n+=1
                            elif int(mark) == 2:
                                if n <= c1:
                                    m2m_1 +=1
                                elif n <= c1 + c2:
                                    m2m_2 += 1
                                elif n <= c1 + c2 + c3:
                                    m2m_3 += 1
                                else:
                                    m2m_4 += 1
                                n+=1
                            elif int(mark) == 1:
                                if n <= c1:
                                    m1m_1 +=1
                                elif n <= c1 + c2:
                                    m1m_2 += 1
                                elif n <= c1 + c2 + c3:
                                    m1m_3 += 1
                                else:
                                    m1m_4 += 1
                                n+=1
                        for row in records:
                            if row[0] == s['Предмет']:
                                m5f, m4f, m3f, m2f, m1f = sum(map(int,row[1].split())), sum(map(int,row[2].split())), sum(map(int,row[3].split())), sum(map(int,row[4].split())), sum(map(int,row[5].split())) #старые
                                p = 'Предмет'
                                if len(s[p]) > 32:
                                    prr = s[p][0:30] + '...'
                                else:
                                    prr = s[p]
                                #ws.add(ws1)
                                if m5m_1 + m5m_2 + m5m_3 +m5m_4 >= m5f:
                                    for i in range(m5m_1 + m5m_2 + m5m_3 +m5m_4 - m5f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        m_id = await create_new_id()
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        cursor.execute("INSERT INTO marks VALUES(?,?,?,?,?);", [m_id, rowrt[0], 5, s[p], d[0:len(d)-7]])
                                        connect.commit()
                                        am+=1
                                        if r1239[1] == 'RU':
                                            ws2 = types.InlineKeyboardButton(text = '👪Поделиться оценкой', switch_inline_query=str(m_id))
                                        elif r1239[1] == 'EN':
                                            ws2 = types.InlineKeyboardButton(text = '👪Share this mark', switch_inline_query=str(m_id))
                                        ws.add(ws2)
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'🟢<b>Новая оценка</b> по предмету <b>"{s[p]}"</b>: 5 🟢'
                                                    f'\n\n'
                                                    fr'Дата выставления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'🟢<b>New mark</b> on the subject<b>"{s[p]}"</b>: 5 🟢'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                else:
                                    for i in range(0 - m5m_1 - m5m_2 - m5m_3 - m5m_4 + m5f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Ваша оценка 5</b> по предмету <b>"{s[p]}"</b> была удалена❎.'
                                                    f'\n\n'
                                                    fr'Дата удаления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Your 5 mark</b> on the subject<b>"{s[p]}"</b> has been deleted❎.'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                if m4m_1 + m4m_2 + m4m_3 +m4m_4 >= m4f:
                                    for i in range(m4m_1 + m4m_2 + m4m_3 +m4m_4 - m4f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        m_id = await create_new_id()
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        cursor.execute("INSERT INTO marks VALUES(?,?,?,?,?);", [m_id, rowrt[0], 4, s[p], d[0:len(d)-7]])
                                        connect.commit()
                                        am+=1
                                        if r1239[1] == 'RU':
                                            ws2 = types.InlineKeyboardButton(text = '👪Поделиться оценкой', switch_inline_query=str(m_id))
                                        elif r1239[1] == 'EN':
                                            ws2 = types.InlineKeyboardButton(text = '👪Share this mark', switch_inline_query=str(m_id))
                                        ws.add(ws2)
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'🔵<b>Новая оценка</b> по предмету <b>"{s[p]}"</b>: 4 🔵'
                                                    f'\n\n'
                                                    fr'Дата выставления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'🔵<b>New mark</b> on the subject<b>"{s[p]}"</b>: 4 🔵'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                else:
                                    for i in range(0 - m4m_1 - m4m_2 - m4m_3 - m4m_4 + m4f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Ваша оценка 4</b> по предмету <b>"{s[p]}"</b> была удалена❎.'
                                                    f'\n\n'
                                                    fr'Дата удаления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Your 4 mark</b> on the subject<b>"{s[p]}"</b> has been deleted❎.'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                if m3m_1 + m3m_2 + m3m_3 +m3m_4 >= m3f:
                                    for i in range(m3m_1 + m3m_2 + m3m_3 +m3m_4 - m3f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        m_id = await create_new_id()
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        cursor.execute("INSERT INTO marks VALUES(?,?,?,?,?);", [m_id, rowrt[0], 3, s[p], d[0:len(d)-7]])
                                        connect.commit()
                                        am+=1
                                        if r1239[1] == 'RU':
                                            ws2 = types.InlineKeyboardButton(text = '👪Поделиться оценкой', switch_inline_query=str(m_id))
                                        elif r1239[1] == 'EN':
                                            ws2 = types.InlineKeyboardButton(text = '👪Share this mark', switch_inline_query=str(m_id))
                                        ws.add(ws2)
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'🟠<b>Новая оценка</b> по предмету <b>"{s[p]}"</b>: 3 🟠'
                                                    f'\n\n'
                                                    fr'Дата выставления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'🟠<b>New mark</b> on the subject<b>"{s[p]}"</b>: 3 🟠'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                else:
                                    for i in range(0 - m3m_1 - m3m_2 - m3m_3 - m3m_4 + m3f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Ваша оценка 3</b> по предмету <b>"{s[p]}"</b> была удалена❎.'
                                                    f'\n\n'
                                                    fr'Дата удаления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Your 3 mark</b> on the subject<b>"{s[p]}"</b> has been deleted❎.'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                if m2m_1 + m2m_2 + m2m_3 +m2m_4 >= m2f:
                                    for i in range(m2m_1 + m2m_2 + m2m_3 +m2m_4 - m2f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        m_id = await create_new_id()
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        cursor.execute("INSERT INTO marks VALUES(?,?,?,?,?);", [m_id, rowrt[0], 2, s[p], d[0:len(d)-7]])
                                        connect.commit()
                                        am+=1
                                        if r1239[1] == 'RU':
                                            ws2 = types.InlineKeyboardButton(text = '👪Поделиться оценкой', switch_inline_query=str(m_id))
                                        elif r1239[1] == 'EN':
                                            ws2 = types.InlineKeyboardButton(text = '👪Share this mark', switch_inline_query=str(m_id))
                                        ws.add(ws2)
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'🔴<b>Новая оценка</b> по предмету <b>"{s[p]}"</b>: 2 🔴'
                                                    f'\n\n'
                                                    fr'Дата выставления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'🔴<b>New mark</b> on the subject<b>"{s[p]}"</b>: 2 🔴'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                else:
                                    for i in range(0 - m2m_1 - m2m_2 - m2m_3 - m2m_4 + m2f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Ваша оценка 2</b> по предмету <b>"{s[p]}"</b> была удалена❎.'
                                                    f'\n\n'
                                                    fr'Дата удаления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Your 2 mark</b> on the subject<b>"{s[p]}"</b> has been deleted❎.'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                if m1m_1 + m1m_2 + m1m_3 + m1m_4 >= m1f:
                                    for i in range(m1m_1 + m1m_2 + m1m_3 + m1m_4 - m1f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        m_id = await create_new_id()
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        cursor.execute("INSERT INTO marks VALUES(?,?,?,?,?);", [m_id, rowrt[0], 1, s[p], d[0:len(d)-7]])
                                        connect.commit()
                                        am+=1
                                        if r1239[1] == 'RU':
                                            ws2 = types.InlineKeyboardButton(text = '👪Поделиться оценкой', switch_inline_query=str(m_id))
                                        elif r1239[1] == 'EN':
                                            ws2 = types.InlineKeyboardButton(text = '👪Share this mark', switch_inline_query=str(m_id))
                                        ws.add(ws2)
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'🔴<b>Новая оценка</b> по предмету <b>"{s[p]}"</b>: 1 🔴'
                                                    f'\n\n'
                                                    fr'Дата выставления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'🔴<b>New mark</b> on the subject<b>"{s[p]}"</b>: 1 🔴'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                else:
                                    for i in range(0 - m1m_1 - m1m_2 - m1m_3 - m1m_4 + m1f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Ваша оценка 1</b> по предмету <b>"{s[p]}"</b> была удалена❎.'
                                                    f'\n\n'
                                                    fr'Дата удаления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Your 1 mark</b> on the subject<b>"{s[p]}"</b> has been deleted❎.'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                        ser = [s['Предмет'], f'{m5m_1} {m5m_2} {m5m_3} {m5m_4}', f'{m4m_1} {m4m_2} {m4m_3} {m4m_4}', f'{m3m_1} {m3m_2} {m3m_3} {m3m_4}', f'{m2m_1} {m2m_2} {m2m_3} {m2m_4}', f'{m1m_1} {m1m_2} {m1m_3} {m1m_4}', s['str']]
                        cursor.execute(f"INSERT INTO t{rowrt[0]} VALUES(?,?,?,?,?,?,?);", ser)
                        connect.commit()
                else:
                    for s in spg:
                        m1m_1, m2m_1, m3m_1, m4m_1, m5m_1, n = 0, 0, 0, 0, 0, 1
                        m1m_2, m2m_2, m3m_2, m4m_2, m5m_2 = 0, 0, 0, 0, 0
                        m1m_3, m2m_3, m3m_3, m4m_3, m5m_3 = 0, 0, 0, 0, 0
                        c1, c2, c3 = map(int, s['Colvo'].split())
                        for mark in list(s['Оценки'].split()):
                            if int(mark) == 5:
                                if n <= c1:
                                    m5m_1 +=1
                                elif n <= c1 + c2:
                                    m5m_2 += 1
                                else:
                                    m5m_3 += 1
                                n+=1
                            elif int(mark) == 4:
                                if n <= c1:
                                    m4m_1 +=1
                                elif n <= c1 + c2:
                                    m4m_2 += 1
                                else:
                                    m4m_3 += 1
                                n+=1
                            elif int(mark) == 3:
                                if n <= c1:
                                    m3m_1 +=1
                                elif n <= c1 + c2:
                                    m3m_2 += 1
                                else:
                                    m3m_3 += 1
                                n+=1
                            elif int(mark) == 2:
                                if n <= c1:
                                    m2m_1 +=1
                                elif n <= c1 + c2:
                                    m2m_2 += 1
                                else:
                                    m2m_3 += 1
                                n+=1
                            elif int(mark) == 1:
                                if n <= c1:
                                    m1m_1 +=1
                                elif n <= c1 + c2:
                                    m1m_2 += 1
                                else:
                                    m1m_3 += 1
                                n+=1
                        for row in records:
                            if row[0] == s['Предмет']:
                                m5f, m4f, m3f, m2f, m1f = sum(map(int,row[1].split())), sum(map(int,row[2].split())), sum(map(int,row[3].split())), sum(map(int,row[4].split())), sum(map(int,row[5].split())) #старые
                                ws = types.InlineKeyboardMarkup()
                                ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                p = 'Предмет'
                                if len(s[p]) > 32:
                                    prr = s[p][0:30] + '...'
                                else:
                                    prr = s[p]
                                ws.add(types.InlineKeyboardButton(text = f'📊Статистика по предмету', callback_data = f'P{prr}'))
                                #ws2 = types.InlineKeyboardMarkup(text = 'Посмотреть статистику по предмету', callback_data='qwertyuiop')
                                ws.add(ws1)
                                if m5m_1 + m5m_2 + m5m_3 >= m5f:
                                    for i in range(m5m_1 + m5m_2 + m5m_3 - m5f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        m_id = await create_new_id()
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        if r1239[1] == 'RU':
                                            ws2 = types.InlineKeyboardButton(text = '👪Поделиться оценкой', switch_inline_query=str(m_id))
                                        elif r1239[1] == 'EN':
                                            ws2 = types.InlineKeyboardButton(text = '👪Share this mark', switch_inline_query=str(m_id))
                                        ws.add(ws2)
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'🟢<b>Новая оценка</b> по предмету <b>"{s[p]}"</b>: 5 🟢'
                                                    f'\n\n'
                                                    fr'Дата выставления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'🟢<b>New mark</b> on the subject<b>"{s[p]}"</b>: 5 🟢'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                else:
                                    for i in range(0 - m5m_1 - m5m_2 - m5m_3 + m5f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Ваша оценка 5</b> по предмету <b>"{s[p]}"</b> была удалена❎.'
                                                    f'\n\n'
                                                    fr'Дата удаления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Your 5 mark</b> on the subject<b>"{s[p]}"</b> has been deleted❎.'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                if m4m_1 + m4m_2 + m4m_3 >= m4f:
                                    for i in range(m4m_1 + m4m_2 + m4m_3 - m4f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        m_id = await create_new_id()
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        if r1239[1] == 'RU':
                                            ws2 = types.InlineKeyboardButton(text = '👪Поделиться оценкой', switch_inline_query=str(m_id))
                                        elif r1239[1] == 'EN':
                                            ws2 = types.InlineKeyboardButton(text = '👪Share this mark', switch_inline_query=str(m_id))
                                        ws.add(ws2)
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'🔵<b>Новая оценка</b> по предмету <b>"{s[p]}"</b>: 4 🔵'
                                                    f'\n\n'
                                                    fr'Дата выставления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'🔵<b>New mark</b> on the subject<b>"{s[p]}"</b>: 4 🔵'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                else:
                                    for i in range(0 - m4m_1 - m4m_2 - m4m_3 + m4f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Ваша оценка 4</b> по предмету <b>"{s[p]}"</b> была удалена❎.'
                                                    f'\n\n'
                                                    fr'Дата удаления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Your 4 mark</b> on the subject<b>"{s[p]}"</b> has been deleted❎.'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                if m3m_1 + m3m_2 + m3m_3 >= m3f:
                                    for i in range(m3m_1 + m3m_2 + m3m_3 - m3f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        m_id = await create_new_id()
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        if r1239[1] == 'RU':
                                            ws2 = types.InlineKeyboardButton(text = '👪Поделиться оценкой', switch_inline_query=str(m_id))
                                        elif r1239[1] == 'EN':
                                            ws2 = types.InlineKeyboardButton(text = '👪Share this mark', switch_inline_query=str(m_id))
                                        ws.add(ws2)
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'🟠<b>Новая оценка</b> по предмету <b>"{s[p]}"</b>: 3 🟠'
                                                    f'\n\n'
                                                    fr'Дата выставления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'🟠<b>New mark</b> on the subject<b>"{s[p]}"</b>: 3 🟠'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                else:
                                    for i in range(0 - m3m_1 - m3m_2 - m3m_3 + m3f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Ваша оценка 3</b> по предмету <b>"{s[p]}"</b> была удалена❎.'
                                                    f'\n\n'
                                                    fr'Дата удаления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Your 3 mark</b> on the subject<b>"{s[p]}"</b> has been deleted❎.'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                if m2m_1 + m2m_2 + m2m_3 >= m2f:
                                    for i in range(m2m_1 + m2m_2 + m2m_3 - m2f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        m_id = await create_new_id()
                                        if r1239[1] == 'RU':
                                            ws2 = types.InlineKeyboardButton(text = '👪Поделиться оценкой', switch_inline_query=str(m_id))
                                        elif r1239[1] == 'EN':
                                            ws2 = types.InlineKeyboardButton(text = '👪Share this mark', switch_inline_query=str(m_id))
                                        ws.add(ws2)
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'🔴<b>Новая оценка</b> по предмету <b>"{s[p]}"</b>: 2 🔴'
                                                    f'\n\n'
                                                    fr'Дата выставления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'🔴<b>New mark</b> on the subject<b>"{s[p]}"</b>: 2 🔴'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                else:
                                    for i in range(0 - m2m_1 - m2m_2 - m2m_3 + m2f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Ваша оценка 2</b> по предмету <b>"{s[p]}"</b> была удалена❎.'
                                                    f'\n\n'
                                                    fr'Дата удаления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Your 2 mark</b> on the subject<b>"{s[p]}"</b> has been deleted❎.'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                if m1m_1 + m1m_2 + m1m_3 >= m1f:
                                    for i in range(m1m_1 + m1m_2 + m1m_3 - m1f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        m_id = await create_new_id()
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        if r1239[1] == 'RU':
                                            ws2 = types.InlineKeyboardButton(text = '👪Поделиться оценкой', switch_inline_query=str(m_id))
                                        elif r1239[1] == 'EN':
                                            ws2 = types.InlineKeyboardButton(text = '👪Share this mark', switch_inline_query=str(m_id))
                                        ws.add(ws2)
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'🔴<b>Новая оценка</b> по предмету <b>"{s[p]}"</b>: 1 🔴'
                                                    f'\n\n'
                                                    fr'Дата выставления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'🔴<b>New mark</b> on the subject<b>"{s[p]}"</b>: 1 🔴'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                                else:
                                    for i in range(0 - m1m_1 - m1m_2 - m1m_3 + m1f):
                                        if r1239[1] == 'RU':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀Посмотреть', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Статистика', callback_data = f'P{prr}'), ws1)
                                        elif r1239[1] == 'EN':
                                            ws = types.InlineKeyboardMarkup()
                                            ws1 = types.InlineKeyboardButton(text = '👀See', url = 'https://elschool.ru/')
                                            ws.add(types.InlineKeyboardButton(text = f'📊Statistics', callback_data = f'P{prr}'), ws1)
                                        p, d = 'Предмет', str(datetime.datetime.now())
                                        am+=1
                                        try:
                                            if r1239[1] == 'RU':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Ваша оценка 1</b> по предмету <b>"{s[p]}"</b> была удалена❎.'
                                                    f'\n\n'
                                                    fr'Дата удаления: <b>{d[0:len(d)-7]} МСК.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                            elif r1239[1] == 'EN':
                                                await bot.send_message(rowrt[0], (fr'❎<b>Your 1 mark</b> on the subject<b>"{s[p]}"</b> has been deleted❎.'
                                                    f'\n\n'
                                                    fr'Date: <b>{d[0:len(d)-7]} MSK.</b>'), reply_markup=ws, parse_mode='HTML')
                                                await asyncio.sleep(1.0)
                                        except:
                                            await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
                                            await asyncio.sleep(1.0)
                        ser = [s['Предмет'], f'{m5m_1} {m5m_2} {m5m_3}', f'{m4m_1} {m4m_2} {m4m_3}', f'{m3m_1} {m3m_2} {m3m_3}', f'{m2m_1} {m2m_2} {m2m_3}', f'{m1m_1} {m1m_2} {m1m_3}', s['str']]
                        cursor.execute(f"INSERT INTO t{rowrt[0]} VALUES(?,?,?,?,?,?,?);", ser)
                        connect.commit()
            await asyncio.sleep(2.0)
        if am == 0 and by_user:
            rfv = types.InlineKeyboardMarkup()
            rfv.add(types.InlineKeyboardButton(text='❎', callback_data='delete'))
            await bot.send_message(user_id, "🔕Новых оценок не обнаружено!", reply_markup=rfv)

        if by_user:
            cursor.execute('DELETE FROM time WHERE user_id=?', (user_id,))
            connect.commit()
            cursor.execute("INSERT INTO time VALUES(?,?,?);", [user_id, int(time.time()), last_time[2]])
            connect.commit()

    except Exception:
        await bot.send_message(-1001984000978, f'bug: {rowrt[0]}\n\n{traceback.format_exc()}')
        await bot.send_message(user_id, "❎ Произошла какая-то ошибка", reply_markup=rfv)
        await asyncio.sleep(1.0)


@bot.callback_query_handler(lambda call: call.data == 'mm')
async def mm(call):
    mm = types.InlineKeyboardMarkup()
    mm.add(types.InlineKeyboardButton(text='📊 Статистика', callback_data='stat1'))
    mm.add(types.InlineKeyboardButton(text='📝 Табель', callback_data='tabel'))
    mm.add(types.InlineKeyboardButton(text='🔎 Проверить новые оценки', callback_data='parse'))
    mm.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='menu'))
    try:
        await bot.edit_message_media(chat_id=call.from_user.id, message_id=call.message.message_id, media=types.InputMediaPhoto(media=IMAGES['marks'], caption='*🔀 Выберите* необходимый пункт *меню оценок*:', parse_mode='markdown'), reply_markup=mm)
    except Exception:
        await bot.send_photo(call.from_user.id, 'https://imgur.com/Xgql7hf.jpg', '*🔀 Выберите* необходимый пункт *меню оценок*:', parse_mode='markdown', reply_markup=mm)
        try:
            await bot.delete_message(call.message.chat.id, call.message.message_id)
        except Exception:
            pass


@bot.callback_query_handler(lambda call: call.data == 'parse')
async def not_parse_func(call):
    lk = types.InlineKeyboardMarkup()
    lk.add(types.InlineKeyboardButton(text='🔙 В меню', callback_data='menu'))
    await bot.send_message(call.from_user.id, '❗*Функция переехала*\n\n💬 Для обновления оценок используйте *команду* /get\_marks', parse_mode='markdown', reply_markup=lk)


@bot.callback_query_handler(lambda call: call.data == 'tabel')
async def tabel(call):
    kn = types.InlineKeyboardMarkup()
    kn.add(types.InlineKeyboardButton(text='📥 Скачать табель', callback_data='tabel_get'))
    kn.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='mm'))
    await bot.send_message(call.from_user.id, '*Данная функция* позволяет раз в *24 часа* получить *табель* своих оценок в формате 📗EXCEL', parse_mode='markdown', reply_markup=kn)


@bot.callback_query_handler(lambda call: call.data == 'tabel_get')
async def tabel_get(call):
    cursor.execute("SELECT * FROM time")
    rec, w = cursor.fetchall(), [0, 0, 0]
    for rowrec in rec:
        if rowrec[0] == call.from_user.id:
            w = rowrec
            break
    now_time = time.time()
    #print(now_time - w[1])
    if now_time - w[2] < 3600*24:
        await bot.answer_callback_query(call.id, f'❎Вы уже использовали эту функцию в последние сутки. Попробуйте через {max(int(24-(now_time - w[2])//3600), 1)} часов', show_alert=True)
        return
    try:
        cursor.execute(f"SELECT * FROM t{call.from_user.id}")
        records = cursor.fetchall()
        from openpyxl import Workbook
        from openpyxl.styles import Border, Side, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        if len(list(records[0][1].split())) == 4:
            ws['B1'] = "Предмет"
            ws['C1'] = "1 четверть"
            ws['D1'] = "Ср.балл"
            ws['E1'] = "2 четверть"
            ws['F1'] = "Ср.балл"
            ws['G1'] = "3 четверть"
            ws['H1'] = "Ср.балл"
            ws['I1'] = "4 четверть"
            ws['J1'] = "Ср.балл"
        elif len(list(records[0][1].split())) == 3:
            ws['B1'] = "Предмет"
            ws['C1'] = "1 триместр"
            ws['D1'] = "Ср.балл"
            ws['E1'] = "2 триместр"
            ws['F1'] = "Ср.балл"
            ws['G1'] = "3 триместр"
            ws['H1'] = "Ср.балл"
        h = 0
        thin = Side(border_style="thin", color="000000")
        font = Font(name='Calibri', size=11, bold=True, italic=False,  vertAlign=None, underline='none', strike=False, color='FF000000')
        ws['B1'].font = font
        ws['C1'].font = font
        ws['D1'].font = font
        ws['E1'].font = font
        ws['F1'].font = font
        ws['G1'].font = font
        ws['H1'].font = font
        ws['I1'].font = font
        ws['J1'].font = font
        #---
        ws['B1'].border = Border(right=thin)
        ws['C1'].border = Border(left=thin)
        ws['D1'].border = Border(right=thin)
        ws['E1'].border = Border(left=thin)
        ws['F1'].border = Border(right=thin)
        ws['G1'].border = Border(left=thin)
        ws['H1'].border = Border(right=thin)
        ws['I1'].border = Border(left=thin)
        ws['J1'].border = Border(right=thin)
        fill5 = PatternFill(fill_type="solid", fgColor="00FF00")
        fill4 = PatternFill(fill_type="solid", fgColor="00BFFF")
        fill3 = PatternFill(fill_type="solid", fgColor="FF8C00")
        fill2 = PatternFill(fill_type="solid", fgColor="CD5C5C")
        for row in records:
            h+=1
            ws[f'A{h+1}'] = h
            ws[f'B{h+1}'] = row[0]
            try:
                ws[f'C{h+1}'].border = Border(left=thin)
                ws[f'D{h+1}'].border = Border(right=thin)
                n1 = int(row[1].split()[0])+int(row[2].split()[0])+int(row[3].split()[0])+int(row[4].split()[0])+int(row[5].split()[0])
                s1 = round((int(row[1].split()[0])*5+int(row[2].split()[0])*4+int(row[3].split()[0])*3+int(row[4].split()[0])*2+int(row[5].split()[0]))/n1, 2)
                ws[f'C{h+1}'] = ' '.join(list(row[6].split())[0:n1])
                ws[f'D{h+1}'] = s1
                if s1 >= 4.5:
                    ws[f'D{h+1}'].fill = fill5
                elif s1 >= 3.5:
                    ws[f'D{h+1}'].fill = fill4
                elif s1 >= 2.5:
                    ws[f'D{h+1}'].fill = fill3
                else:
                    ws[f'D{h+1}'].fill = fill2
            except:
                pass
            try:
                ws[f'E{h+1}'].border = Border(left=thin)
                ws[f'F{h+1}'].border = Border(right=thin)
                n2 = int(row[1].split()[1])+int(row[2].split()[1])+int(row[3].split()[1])+int(row[4].split()[1])+int(row[5].split()[1])
                s2 = round((int(row[1].split()[1])*5+int(row[2].split()[1])*4+int(row[3].split()[1])*3+int(row[4].split()[1])*2+int(row[5].split()[1]))/n2, 2)
                ws[f'E{h+1}'] = ' '.join(list(row[6].split())[n1:n1+n2])
                ws[f'F{h+1}'] = s2
                if s2 >= 4.5:
                    ws[f'F{h+1}'].fill = fill5
                elif s2 >= 3.5:
                    ws[f'F{h+1}'].fill = fill4
                elif s2 >= 2.5:
                    ws[f'F{h+1}'].fill = fill3
                else:
                    ws[f'F{h+1}'].fill = fill2
            except:
                pass
            try:
                ws[f'G{h+1}'].border = Border(left=thin)
                ws[f'H{h+1}'].border = Border(right=thin)
                n3 = int(row[1].split()[2])+int(row[2].split()[2])+int(row[3].split()[2])+int(row[4].split()[2])+int(row[5].split()[2])
                s3 = round((int(row[1].split()[2])*5+int(row[2].split()[2])*4+int(row[3].split()[2])*3+int(row[4].split()[2])*2+int(row[5].split()[2]))/n3, 2)
                ws[f'G{h+1}'] = ' '.join(list(row[6].split())[n1+n2:n1+n2+n3])
                ws[f'H{h+1}'] = s3
                if s3 >= 4.5:
                    ws[f'H{h+1}'].fill = fill5
                elif s3 >= 3.5:
                    ws[f'H{h+1}'].fill = fill4
                elif s3 >= 2.5:
                    ws[f'H{h+1}'].fill = fill3
                else:
                    ws[f'H{h+1}'].fill = fill2
            except:
                pass
            try:
                ws[f'I{h+1}'].border = Border(left=thin)
                ws[f'J{h+1}'].border = Border(right=thin)
                n4 = int(row[1].split()[3])+int(row[2].split()[3])+int(row[3].split()[3])+int(row[4].split()[3])+int(row[5].split()[3])
                s4 = round((int(row[1].split()[3])*5+int(row[2].split()[3])*4+int(row[3].split()[3])*3+int(row[4].split()[3])*2+int(row[5].split()[3]))/n4, 2)
                ws[f'I{h+1}'] = ' '.join(list(row[6].split())[n1+n2+n3:n1+n2+n3+n4])
                ws[f'J{h+1}'] = s4
                if s4 >= 4.5:
                    ws[f'J{h+1}'].fill = fill5
                elif s4 >= 3.5:
                    ws[f'J{h+1}'].fill = fill4
                elif s4 >= 2.5:
                    ws[f'J{h+1}'].fill = fill3
                else:
                    ws[f'J{h+1}'].fill = fill2
            except:
                pass
        wb.save(f'{call.from_user.id}tab.xlsx')
        cursor.execute('DELETE FROM time WHERE user_id=?', (call.from_user.id,))
        connect.commit()
        cursor.execute("INSERT INTO time VALUES(?,?,?);", [call.from_user.id, w[1], int(time.time())])
        connect.commit()
        await bot.send_document(call.from_user.id, open(f'{call.from_user.id}tab.xlsx', 'rb'), caption='📗Табель в формате Excel', visible_file_name='табель.xlsx')
        os.remove(f'{call.from_user.id}tab.xlsx')
    except:
        #print(traceback.format_exc())
        await bot.answer_callback_query(call.id, '❎Произошла ошибка. Попробуйте обновить оценки с помощью команды /get_marks', show_alert=True)
        return

@bot.callback_query_handler(lambda call: call.data[0]=="C")
async def calc(call):
    number = int(call.data[1])
    predm = call.data[2:]
    #print(predm)
    await bot.answer_callback_query(call.id, '🔜', show_alert=True)


@bot.callback_query_handler(lambda call: call.data[0:4] == 'stat')
async def stat(call):
    if call.data[4:] == '':
        n = 1
    else:
        n = int(call.data[4:])
    #mst1 = types.InlineKeyboardButton(text = '👪Поделиться', switch_inline_query = 'potom sdelau')
    zx2 = types.InlineKeyboardButton(text = '🔙 В меню', callback_data = 'menu')
    try:
        cursor.execute(f"SELECT * FROM t{call.from_user.id}")
        records = cursor.fetchall()
        amount_all_marks, sum_all_marks, sp_maxball, sp_minball = 0, 0, [[-1, '']], [[6, '']]
        predm = types.InlineKeyboardMarkup()
        #predm.add(mst1)
        h = 0
        for row in records:
            h+=1
            if h in range(6*(n-1) + 1, 6*n + 1):
                if len(row[0]) > 32:
                    prr = row[0][0:30] + '...'
                else:
                    prr = row[0]
                predm.add(types.InlineKeyboardButton(text = f'{row[0]}', callback_data = f'P{prr}'))
            amount_all_marks += sum(map(int,row[1].split())) * 5
            sum_all_marks += sum(map(int,row[1].split()))
            amount_all_marks += sum(map(int,row[2].split())) * 4
            sum_all_marks += sum(map(int,row[2].split()))
            amount_all_marks += sum(map(int,row[3].split())) * 3
            sum_all_marks += sum(map(int,row[3].split()))
            amount_all_marks += sum(map(int,row[4].split())) * 2
            sum_all_marks += sum(map(int,row[4].split()))
            amount_all_marks += sum(map(int,row[5].split())) * 1
            sum_all_marks += sum(map(int,row[5].split()))

            amount_all_marks12 = sum(map(int,row[1].split())) * 5
            sum_all_marks12 = sum(map(int,row[1].split()))
            amount_all_marks12 += sum(map(int,row[2].split())) * 4
            sum_all_marks12 += sum(map(int,row[2].split()))
            amount_all_marks12 += sum(map(int,row[3].split())) * 3
            sum_all_marks12 += sum(map(int,row[3].split()))
            amount_all_marks12 += sum(map(int,row[4].split())) * 2
            sum_all_marks12 += sum(map(int,row[4].split()))
            amount_all_marks12 += sum(map(int,row[5].split())) * 1
            sum_all_marks12 += sum(map(int,row[5].split()))
            if sum_all_marks12 == 0:
                sum_all_marks12 += 1
            sr_ball = amount_all_marks12 / sum_all_marks12
            if sr_ball > sp_maxball[0][0] and sr_ball != 0:
                sp_maxball = [[sr_ball, row[0]]]
            elif sr_ball == sp_maxball[0][0]:
                sp_maxball.append([sr_ball, row[0]])

            if sr_ball < sp_minball[0][0] and sr_ball != 0:
                sp_minball = [[sr_ball, row[0]]]

            elif sr_ball == sp_minball[0][0]:
                sp_minball.append([sr_ball, row[0]])

        if sum_all_marks == 0:
            sum_all_marks += 1

        strmaxpr, strminpr = '', ''
        for r in sp_maxball:
            strmaxpr += (f'*"{r[1]}"*, ')
        for r in sp_minball:
            strminpr += (f'*"{r[1]}"*, ')
        t = len(records)
        r = 0
        if t % 6 != 0:
            r=1
        if n != int(t/6) + r and n!=1:
            predm.add(types.InlineKeyboardButton(text = '⬅', callback_data=f'stat{n-1}'), types.InlineKeyboardButton(text = '➡', callback_data=f'stat{n+1}'))
        elif n != 1:
            predm.add(types.InlineKeyboardButton(text = '⬅', callback_data=f'stat{n-1}'))
        elif n != int(t/6) + r:
            predm.add(types.InlineKeyboardButton(text = '➡', callback_data=f'stat{n+1}'))
        predm.add(zx2)
        if strmaxpr != '' and strminpr != '' :
            nn = round(amount_all_marks/sum_all_marks, 2)
            nn = str(nn)
            res_s = ''
            for s in nn:
                res_s+=s
            #-----
            nb = round(sp_maxball[0][0], 2)
            nb = str(nb)
            res_b = ''
            for s in nb:
                res_b+=s
            #-----
            nm = round(sp_minball[0][0], 2)
            nm = str(nm)
            res_m = ''
            for s in nm:
                res_m+=s
            n1 = call.from_user.first_name
            if n1 == None:
                n1 = ''
            n2 = call.from_user.last_name
            if n2 == None:
                n2 = ''
            if nb != '-1' and nm != '6':
                await bot.send_photo(call.from_user.id, photo='https://imgur.com/sRyjwZQ.jpg', caption=f'*📊Годовая статистика оценок {n1} {n2}*\n\n🔹{res_s} — *общий* средний балл по *всем* предметам.\n\n🔺{res_b} — *наибольший* средний балл *за год*, по предметам {strmaxpr[0:len(strmaxpr)-2]}.\n\n🔻{res_m} — *наименьший* средний балл *за год*, по предметам {strminpr[0:len(strminpr)-2]}.\n\n_Для просмотра статистики по отдельному предмету используйте кнопки ниже:_', parse_mode='markdown', reply_markup=predm)
            else:
                await bot.send_photo(call.from_user.id, photo='https://imgur.com/sRyjwZQ.jpg', caption=f'*📊Годовая статистика оценок {n1} {n2}*\n\n🔹{res_s} — *общий* средний балл по *всем* предметам.\n\n🔺*Наибольший* средний балл *за год* — _нет данных_.\n\n🔻*Наименьший* средний балл *за год* — _нет данных_.\n\n_Для просмотра статистики по отдельному предмету используйте кнопки ниже:_', parse_mode='markdown', reply_markup=predm)
        else:
            nn = round(amount_all_marks/sum_all_marks, 2)
            nn = str(nn)
            res_s = ''
            for s in nn:
                res_s+=s

            n1 = call.from_user.first_name
            if n1 == None:
                n1 = ''
            n2 = call.from_user.last_name
            if n2 == None:
                n2 = ''
            await bot.send_photo(call.from_user.id, photo='https://imgur.com/sRyjwZQ.jpg', caption=f'*📊Годовая статистика оценок {n1} {n2}*\n\n🔹{res_s} — *общий* средний балл по *всем* предметам.\n\n🔺*Наибольший* средний балл *за год* — _нет данных_.\n\n🔻*Наименьший* средний балл *за год* — _нет данных_.\n\n_Для просмотра статистики по отдельному предмету используйте кнопки ниже:_', parse_mode='markdown', reply_markup=predm)
        try:
            await bot.delete_message(call.from_user.id, call.message.message_id)
        except:
            pass
    except:
        pr = types.InlineKeyboardMarkup()
        pr1 = types.InlineKeyboardButton(text = '➕ Подключить аккаунт ELSCHOOL', callback_data='podkl')
        pr2 = types.InlineKeyboardButton(text = '🔙 Назад', callback_data = 'menu')
        pr.add(pr1)
        pr.add(pr2)
        await bot.send_message(call.from_user.id, '🌐Привяжите аккаунт ELSCHOOL к аккаунту, чтобы получить доступ к разделу 📊Статистика\n\nЕсли вы недавно привязывали новый аккаунт, ожидание обновления данных может занять около 30 минут', reply_markup=pr)
        #await bot.send_message(-1001984000978, f'{traceback.format_exc()}\n\nНе привязан акк')


@bot.callback_query_handler(lambda call: call.data[0] == 'P')
async def predmet(call):
    predmet = call.data[1:]
    cursor.execute(f"SELECT * FROM t{call.from_user.id}")
    records = cursor.fetchall()
    amount_all_marks, sum_all_marks = 0, 0
    for row in records:
        if row[0][0:len(predmet)-3] == predmet[0:len(predmet)-3]:
            summ = 0
            am_marks = 0
            sp_gr = []
            for mark in list(map(int, row[6].split())):
                summ+=mark
                am_marks+=1
                sp_gr.append(round(summ/am_marks, 2))
            import matplotlib.pyplot as plt
            plt.clf()
            fig, ax = plt.subplots()
            plt.title('Изменение среднего балла за год')
            plt.xlabel('количество оценок')
            plt.ylabel('средний балл')
            h = 0
            if len(sp_gr) > 1:
                for mmm in sp_gr:
                    h+=1
                    if h-1:
                        if pr_m >= 4.5:
                            c = 'green'
                        elif pr_m >= 3.5:
                            c = 'blue'
                        elif pr_m >= 2.5:
                            c = 'yellow'
                        else:
                            c = 'red'
                        ax.plot([h-1, h], [pr_m, mmm], linestyle="-",color=c, marker='o')
                    pr_m = mmm
            elif len(sp_gr) == 1:
                pr_m = sp_gr[0]
                if pr_m >= 4.5:
                    c = 'green'
                elif pr_m >= 3.5:
                    c = 'blue'
                elif pr_m >= 2.5:
                    c = 'yellow'
                else:
                    c = 'red'
                ax.plot(1, sp_gr[0], linestyle="-",color=c, marker='o')
            fig.savefig(f'{call.from_user.id}.png')
            plt.close()
            amount_all_marks += sum(map(int,row[1].split())) * 5
            sum_all_marks += sum(map(int,row[1].split()))
            amount_all_marks += sum(map(int,row[2].split())) * 4
            sum_all_marks += sum(map(int,row[2].split()))
            amount_all_marks += sum(map(int,row[3].split())) * 3
            sum_all_marks += sum(map(int,row[3].split()))
            amount_all_marks += sum(map(int,row[4].split())) * 2
            sum_all_marks += sum(map(int,row[4].split()))
            amount_all_marks += sum(map(int,row[5].split())) * 1
            sum_all_marks += sum(map(int,row[5].split()))
            if sum_all_marks == 0:
                sum_all_marks = 1
            sr_ball = amount_all_marks / sum_all_marks
            best, bad = 'нет данных', 'нет данных'
            if sum(map(int, row[5].split())) > 0:
                bad = 1
                best = 1
            if sum(map(int, row[4].split())) > 0:
                if bad == 'нет данных':
                    bad = 2
                best = 2
            if sum(map(int, row[3].split())) > 0:
                if bad == 'нет данных':
                    bad = 3
                best = 3
            if sum(map(int, row[2].split())) > 0:
                if bad == 'нет данных':
                    bad = 4
                best = 4
            if sum(map(int, row[1].split())) > 0:
                if bad == 'нет данных':
                    bad = 5
                best = 5
            pil = types.InlineKeyboardMarkup()
            if len(row[1].split()) == 4:
                pil.add(types.InlineKeyboardButton(text = '1️⃣ четверть', callback_data = f'Q1{row[0][0:30]}'), types.InlineKeyboardButton(text = '2️⃣ четверть', callback_data = f'Q2{row[0][0:30]}'))
                pil.add(types.InlineKeyboardButton(text = '3️⃣ четверть', callback_data = f'Q3{row[0][0:30]}'), types.InlineKeyboardButton(text = '4️⃣ четверть', callback_data = f'Q4{row[0][0:30]}'))
                pil.add(types.InlineKeyboardButton(text = '🔷1 полугодие🔷', callback_data = f'Q5{row[0][0:30]}'))
                pil.add(types.InlineKeyboardButton(text = '🔷2 полугодие🔷', callback_data = f'Q6{row[0][0:30]}'))
                #pil.add(types.InlineKeyboardButton(text = '🆕Цель🎯', callback_data = f'T{row[0][0:30]}'))
            else:
                pil.add(types.InlineKeyboardButton(text = '1️⃣ триместр', callback_data = f'Q1{row[0][0:30]}'))
                pil.add(types.InlineKeyboardButton(text = '2️⃣ триместр', callback_data = f'Q2{row[0][0:30]}'))
                pil.add(types.InlineKeyboardButton(text = '3️⃣ триместр', callback_data = f'Q3{row[0][0:30]}'))
                #pil.add(types.InlineKeyboardButton(text = '🆕Цель🎯', callback_data = f'T{row[0][0:30]}'))
            pil1 = types.InlineKeyboardButton(text = '👪Поделиться', switch_inline_query = row[0])
            piln = types.InlineKeyboardButton(text = '🔙Назад', callback_data = 'stat1')
            #pil.add(pil1)
            pil.add(piln, pil1)
            n1 = call.from_user.first_name
            if n1 == None:
                n1 = ''
            n2 = call.from_user.last_name
            if n2 == None:
                n2 = ''
            await bot.send_photo(call.from_user.id, photo = open(f'{call.from_user.id}.png', 'rb') , caption = f'*📊Годовая статистика оценок {n1} {n2}* по предмету *"{row[0]}"*\n\n🔹*Общий* средний балл *за год* — {round(sr_ball, 2)}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)}, из них:*\n🟢5 — {sum(map(int, row[1].split()))},\n🔵4 — {sum(map(int, row[2].split()))},\n🟠3 — {sum(map(int, row[3].split()))},\n🔴2 — {sum(map(int, row[4].split()))},\n🔴1 — {sum(map(int, row[5].split()))} \n\n_Используйте кнопки ниже для просмотра статистики по отдельным четвертям / триместрам._', parse_mode='markdown', reply_markup=pil)
            os.remove(f'{call.from_user.id}.png')
            break
        #else:
            #print(row[0][0:len(predmet)], predmet)


@bot.callback_query_handler(lambda call: call.data[0] == 'Q')
async def chetv(call):
    #await bot.send_message(call.from_user.id, 'В разработке (доступно только администраторам)')
    #return
    try:
        number = int(call.data[1])
        predm = call.data[2:]
        cursor.execute(f"SELECT * FROM t{call.from_user.id}")
        records = cursor.fetchall()
        for row in records:
            if row[0][0:len(predm)] == predm:
                if number == 1:
                    s, sne= 0, 0
                    s += int(list(row[1].split())[0])
                    s += int(list(row[2].split())[0])
                    s += int(list(row[3].split())[0])
                    s += int(list(row[4].split())[0])
                    s += int(list(row[5].split())[0])
                elif number == 2:
                    s, sne = 0, 0
                    sne += int(list(row[1].split())[0])
                    sne += int(list(row[2].split())[0])
                    sne += int(list(row[3].split())[0])
                    sne += int(list(row[4].split())[0])
                    sne += int(list(row[5].split())[0])
                    s += int(list(row[1].split())[1])
                    s += int(list(row[2].split())[1])
                    s += int(list(row[3].split())[1])
                    s += int(list(row[4].split())[1])
                    s += int(list(row[5].split())[1])
                elif number == 3:
                    s, sne = 0, 0
                    sne += int(list(row[1].split())[0])
                    sne += int(list(row[2].split())[0])
                    sne += int(list(row[3].split())[0])
                    sne += int(list(row[4].split())[0])
                    sne += int(list(row[5].split())[0])
                    sne += int(list(row[1].split())[1])
                    sne += int(list(row[2].split())[1])
                    sne += int(list(row[3].split())[1])
                    sne += int(list(row[4].split())[1])
                    sne += int(list(row[5].split())[1])
                    s += int(list(row[1].split())[2])
                    s += int(list(row[2].split())[2])
                    s += int(list(row[3].split())[2])
                    s += int(list(row[4].split())[2])
                    s += int(list(row[5].split())[2])
                elif number == 4:
                    #s - количество оценок, sne - сколько оценок с начала строки надо пропустить
                    s, sne = 0, 0
                    sne += int(list(row[1].split())[0])
                    sne += int(list(row[2].split())[0])
                    sne += int(list(row[3].split())[0])
                    sne += int(list(row[4].split())[0])
                    sne += int(list(row[5].split())[0])
                    sne += int(list(row[1].split())[1])
                    sne += int(list(row[2].split())[1])
                    sne += int(list(row[3].split())[1])
                    sne += int(list(row[4].split())[1])
                    sne += int(list(row[5].split())[1])
                    sne += int(list(row[1].split())[2])
                    sne += int(list(row[2].split())[2])
                    sne += int(list(row[3].split())[2])
                    sne += int(list(row[4].split())[2])
                    sne += int(list(row[5].split())[2])
                    s += int(list(row[1].split())[3])
                    s += int(list(row[2].split())[3])
                    s += int(list(row[3].split())[3])
                    s += int(list(row[4].split())[3])
                    s += int(list(row[5].split())[3])
                elif number == 5:
                    s, sne = 0, 0
                    s += int(list(row[1].split())[0])
                    s += int(list(row[2].split())[0])
                    s += int(list(row[3].split())[0])
                    s += int(list(row[4].split())[0])
                    s += int(list(row[5].split())[0])
                    s += int(list(row[1].split())[1])
                    s += int(list(row[2].split())[1])
                    s += int(list(row[3].split())[1])
                    s += int(list(row[4].split())[1])
                    s += int(list(row[5].split())[1])
                elif number == 6:
                    s, sne = 0, 0
                    sne += int(list(row[1].split())[0])
                    sne += int(list(row[2].split())[0])
                    sne += int(list(row[3].split())[0])
                    sne += int(list(row[4].split())[0])
                    sne += int(list(row[5].split())[0])
                    sne += int(list(row[1].split())[1])
                    sne += int(list(row[2].split())[1])
                    sne += int(list(row[3].split())[1])
                    sne += int(list(row[4].split())[1])
                    sne += int(list(row[5].split())[1])
                    s += int(list(row[1].split())[2])
                    s += int(list(row[2].split())[2])
                    s += int(list(row[3].split())[2])
                    s += int(list(row[4].split())[2])
                    s += int(list(row[5].split())[2])
                    s += int(list(row[1].split())[3])
                    s += int(list(row[2].split())[3])
                    s += int(list(row[3].split())[3])
                    s += int(list(row[4].split())[3])
                    s += int(list(row[5].split())[3])

                summ, am_marks, sp_gr = 0, 0, []
                if s != 0:
                    sp = list(row[6].split())[sne:sne+s]
                else:
                    sp = []
                for mark in sp:
                    summ+=int(mark)
                    am_marks+=1
                    sp_gr.append(round(summ/am_marks, 2))
                import matplotlib.pyplot as plt
                plt.clf()
                fig, ax = plt.subplots()
                if number < 4:
                    plt.title(f'Изменение среднего балла за {number} четверть/триместр')
                elif number == 4:
                    plt.title(f'Изменение среднего балла за {number} четверть')
                else:
                    plt.title(f'Изменение среднего балла за {number % 4} полугодие')
                plt.xlabel('количество оценок')
                plt.ylabel('средний балл')
                h = 0
                if len(sp_gr) > 1:
                    for mmm in sp_gr:
                        h+=1
                        if h-1:
                            if pr_m >= 4.5:
                                c = 'green'
                            elif pr_m >= 3.5:
                                c = 'blue'
                            elif pr_m >= 2.5:
                                c = 'yellow'
                            else:
                                c = 'red'
                            ax.plot([h-1, h], [pr_m, mmm], linestyle="-",color=c, marker='o')
                        pr_m = mmm
                elif len(sp_gr) == 1:
                    pr_m = sp_gr[0]
                    if pr_m >= 4.5:
                        c = 'green'
                    elif pr_m >= 3.5:
                        c = 'blue'
                    elif pr_m >= 2.5:
                        c = 'yellow'
                    else:
                        c = 'red'
                    ax.plot(1, sp_gr[0], linestyle="-",color=c, marker='o')
                fig.savefig(f'{call.from_user.id}.png')
                pil = types.InlineKeyboardMarkup()
                piln = types.InlineKeyboardButton(text = '🔙 Назад', callback_data = 'stat1')
                if number <= 4:
                    pil.add(types.InlineKeyboardButton(text = '🧮Калькулятор', callback_data=f'C{number}{predm}'))
                pil.add(piln, types.InlineKeyboardButton(text = '👪Поделиться', switch_inline_query=f'{predm} {number}'))
                f = 0
                if sp_gr == []:
                    sp_gr = [0.0]
                    f = 1
                    best, bad = 'нет данных', 'нет данных'
                else:
                    best, bad = max(sp),  min(sp)
                n1 = call.from_user.first_name
                if n1 == None:
                    n1 = ''
                n2 = call.from_user.last_name
                if n2 == None:
                    n2 = ''
                if number < 4:
                    await bot.send_photo(call.from_user.id, photo = open(f'{call.from_user.id}.png', 'rb') , caption = f'*📊Cтатистика оценок {n1} {n2}* по предмету "*{row[0]}*" за *{number} четверть/триместр*.\n\n🔹Средний балл  — {sp_gr[-1]}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)-f}, из них:*\n🟢5 — {int(list(row[1].split())[number-1])},\n🔵4 — {int(list(row[2].split())[number-1])},\n🟠3 — {int(list(row[3].split())[number-1])},\n🔴2 — {int(list(row[4].split())[number-1])},\n🔴1 — {int(list(row[5].split())[number-1])}', parse_mode='markdown', reply_markup=pil)
                elif number == 4:
                    await bot.send_photo(call.from_user.id, photo = open(f'{call.from_user.id}.png', 'rb') , caption = f'*📊Cтатистика оценок {n1} {n2}* по предмету "*{row[0]}*" за *{number} четверть*.\n\n🔹Средний балл  — {sp_gr[-1]}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)-f}, из них:*\n🟢5 — {int(list(row[1].split())[number-1])},\n🔵4 — {int(list(row[2].split())[number-1])},\n🟠3 — {int(list(row[3].split())[number-1])},\n🔴2 — {int(list(row[4].split())[number-1])},\n🔴1 — {int(list(row[5].split())[number-1])}', parse_mode='markdown', reply_markup=pil)
                else:
                    if number == 5:
                        await bot.send_photo(call.from_user.id, photo = open(f'{call.from_user.id}.png', 'rb') , caption = f'*📊Cтатистика оценок {n1} {n2}* по предмету "*{row[0]}*" за *{number % 4} полугодие*.\n\n🔹Средний балл  — {sp_gr[-1]}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)-f}, из них:*\n🟢5 — {int(list(row[1].split())[0]) + int(list(row[1].split())[1])},\n🔵4 — {int(list(row[2].split())[0]) + int(list(row[2].split())[1])},\n🟠3 — {int(list(row[3].split())[0]) + int(list(row[3].split())[1])},\n🔴2 — {int(list(row[4].split())[0]) + int(list(row[4].split())[1])},\n🔴1 — {int(list(row[5].split())[0]) + int(list(row[5].split())[1])}', parse_mode='markdown', reply_markup=pil)
                    elif number == 6:
                        await bot.send_photo(call.from_user.id, photo = open(f'{call.from_user.id}.png', 'rb') , caption = f'*📊Cтатистика оценок {n1} {n2}* по предмету "*{row[0]}*" за *{number % 4} полугодие*.\n\n🔹Средний балл  — {sp_gr[-1]}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)-f}, из них:*\n🟢5 — {int(list(row[1].split())[2]) + int(list(row[1].split())[3])},\n🔵4 — {int(list(row[2].split())[2]) + int(list(row[2].split())[3])},\n🟠3 — {int(list(row[3].split())[2]) + int(list(row[3].split())[3])},\n🔴2 — {int(list(row[4].split())[2]) + int(list(row[4].split())[3])},\n🔴1 — {int(list(row[5].split())[2]) + int(list(row[5].split())[3])}', parse_mode='markdown', reply_markup=pil)
                os.remove(f'{call.from_user.id}.png')
                break
    except:
        await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
        await asyncio.sleep(1.0)


@bot.callback_query_handler(lambda call: call.data == 'help')
async def help(call):
    Faq_app = types.WebAppInfo("https://teletype.in/@the_sloth_bear/faq_elshelp")
    faq_button = types.InlineKeyboardButton(text="❔FAQ", web_app=Faq_app)
    qws = types.InlineKeyboardMarkup()
    piln = types.InlineKeyboardButton(text = '🔙 Назад', callback_data = 'menu')
    # qws.add(faq_button)
    qws.add(piln)
    await bot.send_message(call.from_user.id, '*Нашли баг❓ \nЕсть идея или предложение❓*\n\nВы всегда можете обратиться к *@the_sloth_bear*, либо к [специальному боту](https://t.me/elschool_help_support_bot)', parse_mode = 'markdown', reply_markup=qws)


@bot.inline_handler(func=lambda query: len(query.query) > 0)
async def query_text(query):
    flag_is_id = await is_mark_id(query.query)
    if flag_is_id:
        g = int(query.query)
        cursor.execute("SELECT * FROM marks")
        records, flag = cursor.fetchall(), False
        for row in records:
            if int(row[0]) == g and row[1] == query.from_user.id:
                r = row
                flag = True
                break
        if flag:
            n1 = query.from_user.first_name
            if n1 == None:
                n1 = ''
            n2 = query.from_user.last_name
            if n2 == None:
                n2 = ''
            if r[2] == 5:
                marker = '🟢'
            elif r[2] == 4:
                marker = '🔵'
            elif r[2] == 3:
                marker = '🟠'
            else:
                marker = '🔴'
            r = types.InlineQueryResultArticle(id='1', title='👪Поделиться оценкой', description=f'Отправить свою оценку {r[2]} по предмету "{r[3]}" в 💬чат', thumbnail_url='https://imgur.com/weGHPa6.jpg',
                input_message_content = types.InputTextMessageContent(message_text=f'👤*{n1} {n2}* делится своей оценкой по предмету *"{r[3]}"*!\n\n{marker}Оценка: {r[2]}{marker}\n\n✔Получена: *{r[4]} МСК*', parse_mode='markdown'))
            await bot.answer_inline_query(query.id, [r])
        else:
            pass
    else:
        if query.query.split(' ')[-1] in ['1', '2', '3', '4', '5', '6']:
            hj = query.query.split(' ')
            number = int(query.query.split(' ')[-1])
            predm = ' '.join(hj[0:len(hj)-1])
            cursor.execute(f"SELECT * FROM t{query.from_user.id}")
            records = cursor.fetchall()
            for row in records:
                if row[0][0:len(predm)].lower() == predm.lower() and (number != 3 or len(row[1].split()) == 3):
                    ghta = types.InlineKeyboardMarkup()
                    ghta.add(types.InlineKeyboardButton(text='Перейти к 🤖боту', url='t.me/elschool_help_bot'))
                    #await bot.send_photo(call.from_user.id, photo = open(f'{call.from_user.id}.png', 'rb') , caption = f'*📊Годовая статистика оценок {n1} {n2}* по предмету *"{row[0]}"*\n\n🔹*Общий* средний балл *за год* — {round(sr_ball, 2)}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)}, из них:*\n🟢5 — {sum(map(int, row[1].split()))},\n🔵4 — {sum(map(int, row[2].split()))},\n🟠3 — {sum(map(int, row[3].split()))},\n🔴2 — {sum(map(int, row[4].split()))},\n🔴1 — {sum(map(int, row[5].split()))} \n\n_Используйте кнопки ниже для просмотра статистики по отдельным четвертям / триместрам._', parse_mode='markdown', reply_markup=pil)
                    r = types.InlineQueryResultArticle(id='3', title='👪Поделиться статистикой', description=f'🧩Отправить свою статистику по предмету "{row[0]}" в 💬чат', thumbnail_url='https://imgur.com/weGHPa6.jpg', reply_markup=ghta,
                        input_message_content=types.InputTextMessageContent(message_text=f'Загрузка сообщения...\nЭто может занять некоторое время', parse_mode='markdown'))
                        #input_message_content = types.InputMediaPhoto('https://imgur.com/weGHPa6.jpg', caption='ok'))
                    await bot.answer_inline_query(query.id, [r], cache_time=0)
                    break



        else:
            predm = query.query
            cursor.execute(f"""CREATE TABLE IF NOT EXISTS t{query.from_user.id}(
                predmet TEXT,
                m5_marks TEXT,
                m4_marks TEXT,
                m3_marks TEXT,
                m2_marks TEXT,
                m1_marks TEXT,
                str_marks TEXT
                )
                """)
            connect.commit()
            cursor.execute(f"SELECT * FROM t{query.from_user.id}")
            records, flag_t = cursor.fetchall(), True
            for row in records:
                if row[0].lower() == predm.lower():
                    ghta = types.InlineKeyboardMarkup()
                    ghta.add(types.InlineKeyboardButton(text='Перейти к 🤖боту', url='t.me/elschool_help_bot'))
                    #await bot.send_photo(call.from_user.id, photo = open(f'{call.from_user.id}.png', 'rb') , caption = f'*📊Годовая статистика оценок {n1} {n2}* по предмету *"{row[0]}"*\n\n🔹*Общий* средний балл *за год* — {round(sr_ball, 2)}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)}, из них:*\n🟢5 — {sum(map(int, row[1].split()))},\n🔵4 — {sum(map(int, row[2].split()))},\n🟠3 — {sum(map(int, row[3].split()))},\n🔴2 — {sum(map(int, row[4].split()))},\n🔴1 — {sum(map(int, row[5].split()))} \n\n_Используйте кнопки ниже для просмотра статистики по отдельным четвертям / триместрам._', parse_mode='markdown', reply_markup=pil)
                    r = types.InlineQueryResultArticle(id='2', title='👪Поделиться статистикой', description=f'Отправить свою статистику по предмету "{row[0]}" в 💬чат', thumbnail_url='https://imgur.com/weGHPa6.jpg', reply_markup=ghta,
                        input_message_content=types.InputTextMessageContent(message_text=f'Загрузка сообщения...\nЭто может занять некоторое время', parse_mode='markdown'))
                        #input_message_content = types.InputMediaPhoto('https://imgur.com/weGHPa6.jpg', caption='ok'))
                    await bot.answer_inline_query(query.id, [r], cache_time=0)
                    flag_t = False
                    break
            if flag_t:
                #r = types.InlineQueryResultsButton('Предмет / оценка не найдены')
                #await bot.answer_inline_query(query.id, is_personal=True, cache_time=0, switch_pm_text='опаопаопа')
                pass

async def is_mark_id(m_id):
    s = '0123456789'
    for t in m_id:
        if not t in s:
            return False
    return True


@bot.chosen_inline_handler(func = lambda inline_query: inline_query.result_id == '2')
async def chosen_in_handler(inline_query):
    predmet = inline_query.query
    cursor.execute(f"SELECT * FROM t{inline_query.from_user.id}")
    records = cursor.fetchall()
    amount_all_marks, sum_all_marks = 0, 0
    for row in records:
        if row[0].lower() == predmet.lower():
            summ = 0
            am_marks = 0
            sp_gr = []
            for mark in list(map(int, row[6].split())):
                summ+=mark
                am_marks+=1
                sp_gr.append(round(summ/am_marks, 2))
            import matplotlib.pyplot as plt
            plt.clf()
            fig, ax = plt.subplots()
            plt.title('Изменение среднего балла за год')
            plt.xlabel('количество оценок')
            plt.ylabel('средний балл')
            h = 0
            if len(sp_gr) > 1:
                for mmm in sp_gr:
                    h+=1
                    if h-1:
                        if pr_m >= 4.5:
                            c = 'green'
                        elif pr_m >= 3.5:
                            c = 'blue'
                        elif pr_m >= 2.5:
                            c = 'yellow'
                        else:
                            c = 'red'
                        ax.plot([h-1, h], [pr_m, mmm], linestyle="-",color=c, marker='o')
                    pr_m = mmm
            elif len(sp_gr) == 1:
                pr_m = sp_gr[0]
                if pr_m >= 4.5:
                    c = 'green'
                elif pr_m >= 3.5:
                    c = 'blue'
                elif pr_m >= 2.5:
                    c = 'yellow'
                else:
                    c = 'red'
                ax.plot(1, sp_gr[0], linestyle="-",color=c, marker='o')
            fig.savefig(f'{inline_query.from_user.id}.png')
            amount_all_marks += sum(map(int,row[1].split())) * 5
            sum_all_marks += sum(map(int,row[1].split()))
            amount_all_marks += sum(map(int,row[2].split())) * 4
            sum_all_marks += sum(map(int,row[2].split()))
            amount_all_marks += sum(map(int,row[3].split())) * 3
            sum_all_marks += sum(map(int,row[3].split()))
            amount_all_marks += sum(map(int,row[4].split())) * 2
            sum_all_marks += sum(map(int,row[4].split()))
            amount_all_marks += sum(map(int,row[5].split())) * 1
            sum_all_marks += sum(map(int,row[5].split()))
            if sum_all_marks == 0:
                sum_all_marks = 1
            sr_ball = amount_all_marks / sum_all_marks
            best, bad = 'нет данных', 'нет данных'
            if sum(map(int, row[5].split())) > 0:
                bad = 1
                best = 1
            if sum(map(int, row[4].split())) > 0:
                if bad == 'нет данных':
                    bad = 2
                best = 2
            if sum(map(int, row[3].split())) > 0:
                if bad == 'нет данных':
                    bad = 3
                best = 3
            if sum(map(int, row[2].split())) > 0:
                if bad == 'нет данных':
                    bad = 4
                best = 4
            if sum(map(int, row[1].split())) > 0:
                if bad == 'нет данных':
                    bad = 5
                best = 5
            n1 = inline_query.from_user.first_name
            if n1 == None:
                n1 = ''
            n2 = inline_query.from_user.last_name
            if n2 == None:
                n2 = ''

            with open(f'{inline_query.from_user.id}.png', 'rb') as f:
                sp = requests.post(
                    'https://telegra.ph/upload',
                    files={'file': ('file', f, 'image/png')}  # image/gif, image/jpeg, image/jpg, image/png, video/mp4
                ).json()
                url = 'https://telegra.ph' + sp[0]['src']

            #await bot.send_photo(call.from_user.id, photo = open(f'{call.from_user.id}.png', 'rb') , caption = f'*📊Годовая статистика оценок {n1} {n2}* по предмету *"{row[0]}"*\n\n🔹*Общий* средний балл *за год* — {round(sr_ball, 2)}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)}, из них:*\n🟢5 — {sum(map(int, row[1].split()))},\n🔵4 — {sum(map(int, row[2].split()))},\n🟠3 — {sum(map(int, row[3].split()))},\n🔴2 — {sum(map(int, row[4].split()))},\n🔴1 — {sum(map(int, row[5].split()))} \n\n_Используйте кнопки ниже для просмотра статистики по отдельным четвертям / триместрам._', parse_mode='markdown', reply_markup=pil)
            await bot.edit_message_text(inline_message_id = inline_query.inline_message_id, text=f'[ ]({url})\n*👤{n1} {n2}* делится *📊годовой статистикой оценок * по предмету *"{row[0]}"*\n\n🔹*Общий* средний балл *за год* — {round(sr_ball, 2)}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)}, из них:*\n🟢5 — {sum(map(int, row[1].split()))},\n🔵4 — {sum(map(int, row[2].split()))},\n🟠3 — {sum(map(int, row[3].split()))},\n🔴2 — {sum(map(int, row[4].split()))},\n🔴1 — {sum(map(int, row[5].split()))}', parse_mode='markdown')
            break

@bot.chosen_inline_handler(func = lambda inline_query: inline_query.result_id == '3')
async def chosen_in_handler_3(inline_query):
    hj = inline_query.query.split(' ')
    number = int(inline_query.query.split(' ')[-1])
    predm = ' '.join(hj[0:len(hj)-1])
    try:
        cursor.execute(f"SELECT * FROM t{inline_query.from_user.id}")
        records = cursor.fetchall()
        for row in records:
            if row[0][0:len(predm)].lower() == predm.lower():
                if number == 1:
                    s, sne= 0, 0
                    s += int(list(row[1].split())[0])
                    s += int(list(row[2].split())[0])
                    s += int(list(row[3].split())[0])
                    s += int(list(row[4].split())[0])
                    s += int(list(row[5].split())[0])
                elif number == 2:
                    s, sne = 0, 0
                    sne += int(list(row[1].split())[0])
                    sne += int(list(row[2].split())[0])
                    sne += int(list(row[3].split())[0])
                    sne += int(list(row[4].split())[0])
                    sne += int(list(row[5].split())[0])
                    s += int(list(row[1].split())[1])
                    s += int(list(row[2].split())[1])
                    s += int(list(row[3].split())[1])
                    s += int(list(row[4].split())[1])
                    s += int(list(row[5].split())[1])
                elif number == 3:
                    s, sne = 0, 0
                    sne += int(list(row[1].split())[0])
                    sne += int(list(row[2].split())[0])
                    sne += int(list(row[3].split())[0])
                    sne += int(list(row[4].split())[0])
                    sne += int(list(row[5].split())[0])
                    sne += int(list(row[1].split())[1])
                    sne += int(list(row[2].split())[1])
                    sne += int(list(row[3].split())[1])
                    sne += int(list(row[4].split())[1])
                    sne += int(list(row[5].split())[1])
                    s += int(list(row[1].split())[2])
                    s += int(list(row[2].split())[2])
                    s += int(list(row[3].split())[2])
                    s += int(list(row[4].split())[2])
                    s += int(list(row[5].split())[2])
                elif number == 4:
                    #s - количество оценок, sne - сколько оценок с начала строки надо пропустить
                    s, sne = 0, 0
                    sne += int(list(row[1].split())[0])
                    sne += int(list(row[2].split())[0])
                    sne += int(list(row[3].split())[0])
                    sne += int(list(row[4].split())[0])
                    sne += int(list(row[5].split())[0])
                    sne += int(list(row[1].split())[1])
                    sne += int(list(row[2].split())[1])
                    sne += int(list(row[3].split())[1])
                    sne += int(list(row[4].split())[1])
                    sne += int(list(row[5].split())[1])
                    sne += int(list(row[1].split())[2])
                    sne += int(list(row[2].split())[2])
                    sne += int(list(row[3].split())[2])
                    sne += int(list(row[4].split())[2])
                    sne += int(list(row[5].split())[2])
                    s += int(list(row[1].split())[3])
                    s += int(list(row[2].split())[3])
                    s += int(list(row[3].split())[3])
                    s += int(list(row[4].split())[3])
                    s += int(list(row[5].split())[3])
                elif number == 5:
                    s, sne = 0, 0
                    s += int(list(row[1].split())[0])
                    s += int(list(row[2].split())[0])
                    s += int(list(row[3].split())[0])
                    s += int(list(row[4].split())[0])
                    s += int(list(row[5].split())[0])
                    s += int(list(row[1].split())[1])
                    s += int(list(row[2].split())[1])
                    s += int(list(row[3].split())[1])
                    s += int(list(row[4].split())[1])
                    s += int(list(row[5].split())[1])
                elif number == 6:
                    s, sne = 0, 0
                    sne += int(list(row[1].split())[0])
                    sne += int(list(row[2].split())[0])
                    sne += int(list(row[3].split())[0])
                    sne += int(list(row[4].split())[0])
                    sne += int(list(row[5].split())[0])
                    sne += int(list(row[1].split())[1])
                    sne += int(list(row[2].split())[1])
                    sne += int(list(row[3].split())[1])
                    sne += int(list(row[4].split())[1])
                    sne += int(list(row[5].split())[1])
                    s += int(list(row[1].split())[2])
                    s += int(list(row[2].split())[2])
                    s += int(list(row[3].split())[2])
                    s += int(list(row[4].split())[2])
                    s += int(list(row[5].split())[2])
                    s += int(list(row[1].split())[3])
                    s += int(list(row[2].split())[3])
                    s += int(list(row[3].split())[3])
                    s += int(list(row[4].split())[3])
                    s += int(list(row[5].split())[3])

                summ, am_marks, sp_gr = 0, 0, []
                if s != 0:
                    sp = list(row[6].split())[sne:sne+s]
                else:
                    sp = []
                for mark in sp:
                    summ+=int(mark)
                    am_marks+=1
                    sp_gr.append(round(summ/am_marks, 2))
                import matplotlib.pyplot as plt
                plt.clf()
                fig, ax = plt.subplots()
                if number < 4:
                    plt.title(f'Изменение среднего балла за {number} четверть/триместр')
                elif number == 4:
                    plt.title(f'Изменение среднего балла за {number} четверть')
                else:
                    plt.title(f'Изменение среднего балла за {number % 4} полугодие')
                plt.xlabel('количество оценок')
                plt.ylabel('средний балл')
                h = 0
                if len(sp_gr) > 1:
                    for mmm in sp_gr:
                        h+=1
                        if h-1:
                            if pr_m >= 4.5:
                                c = 'green'
                            elif pr_m >= 3.5:
                                c = 'blue'
                            elif pr_m >= 2.5:
                                c = 'yellow'
                            else:
                                c = 'red'
                            ax.plot([h-1, h], [pr_m, mmm], linestyle="-",color=c, marker='o')
                        pr_m = mmm
                elif len(sp_gr) == 1:
                    pr_m = sp_gr[0]
                    if pr_m >= 4.5:
                        c = 'green'
                    elif pr_m >= 3.5:
                        c = 'blue'
                    elif pr_m >= 2.5:
                        c = 'yellow'
                    else:
                        c = 'red'
                    ax.plot(1, sp_gr[0], linestyle="-",color=c, marker='o')
                fig.savefig(f'{inline_query.from_user.id}.png')
                plt.close()
                f = 0
                if sp_gr == []:
                    sp_gr = [0.0]
                    f = 1
                    best, bad = 'нет данных', 'нет данных'
                else:
                    best, bad = max(sp),  min(sp)
                n1 = inline_query.from_user.first_name
                if n1 == None:
                    n1 = ''
                n2 = inline_query.from_user.last_name
                if n2 == None:
                    n2 = ''

                with open(f'{inline_query.from_user.id}.png', 'rb') as fil:
                    sp = requests.post(
                        'https://telegra.ph/upload',
                        files={'file': ('file', fil, 'image/png')}  # image/gif, image/jpeg, image/jpg, image/png, video/mp4
                    ).json()
                    url = 'https://telegra.ph' + sp[0]['src']
                os.remove(f'{inline_query.from_user.id}.png')
                if number < 4:
                    await bot.edit_message_text(inline_message_id=inline_query.inline_message_id, text = f'[ ]({url})*👤{n1} {n2} делится 📊статистикой оценок* по предмету *"{row[0]}"* за *{number} четверть/триместр*.\n\n🔹Средний балл  — {sp_gr[-1]}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)-f}, из них:*\n🟢5 — {int(list(row[1].split())[number-1])},\n🔵4 — {int(list(row[2].split())[number-1])},\n🟠3 — {int(list(row[3].split())[number-1])},\n🔴2 — {int(list(row[4].split())[number-1])},\n🔴1 — {int(list(row[5].split())[number-1])}', parse_mode='markdown')
                elif number == 4:
                    await bot.edit_message_text(inline_message_id=inline_query.inline_message_id, text = f'[ ]({url})*👤{n1} {n2} делится 📊статистикой оценок* по предмету *"{row[0]}"* за *{number} четверть*.\n\n🔹Средний балл  — {sp_gr[-1]}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)-f}, из них:*\n🟢5 — {int(list(row[1].split())[number-1])},\n🔵4 — {int(list(row[2].split())[number-1])},\n🟠3 — {int(list(row[3].split())[number-1])},\n🔴2 — {int(list(row[4].split())[number-1])},\n🔴1 — {int(list(row[5].split())[number-1])}', parse_mode='markdown')
                else:
                    if number == 5:
                        await bot.edit_message_text(inline_message_id=inline_query.inline_message_id, text = f'[ ]({url})*👤{n1} {n2} делится 📊статистикой оценок* по предмету *"{row[0]}"* за *{number % 4} полугодие*.\n\n🔹Средний балл  — {sp_gr[-1]}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)-f}, из них:*\n🟢5 — {int(list(row[1].split())[0]) + int(list(row[1].split())[1])},\n🔵4 — {int(list(row[2].split())[0]) + int(list(row[2].split())[1])},\n🟠3 — {int(list(row[3].split())[0]) + int(list(row[3].split())[1])},\n🔴2 — {int(list(row[4].split())[0]) + int(list(row[4].split())[1])},\n🔴1 — {int(list(row[5].split())[0]) + int(list(row[5].split())[1])}', parse_mode='markdown')
                    elif number == 6:
                        await bot.edit_message_text(inline_message_id=inline_query.inline_message_id, text = f'[ ]({url})*👤{n1} {n2} делится 📊статистикой оценок* по предмету *"{row[0]}"* за *{number % 4} полугодие*.\n\n🔹Средний балл  — {sp_gr[-1]}\n\n*🔝Лучшая* оценка — *{best}*,\n*🔻Худшая* оценка — *{bad}*\n\n*Всего оценок — {len(sp_gr)-f}, из них:*\n🟢5 — {int(list(row[1].split())[2]) + int(list(row[1].split())[3])},\n🔵4 — {int(list(row[2].split())[2]) + int(list(row[2].split())[3])},\n🟠3 — {int(list(row[3].split())[2]) + int(list(row[3].split())[3])},\n🔴2 — {int(list(row[4].split())[2]) + int(list(row[4].split())[3])},\n🔴1 — {int(list(row[5].split())[2]) + int(list(row[5].split())[3])}', parse_mode='markdown')
                break
    except:
        await bot.send_message(-1001984000978, f'{traceback.format_exc()}')
        await asyncio.sleep(1.0)


@bot.callback_query_handler(lambda call: call.data == 'dz')
async def dz(call):
    try:
        await bot.delete_message(call.message.chat.id, call.message.message_id)
    except:
        pass
    days = types.InlineKeyboardMarkup()
    days.add(types.InlineKeyboardButton(text='🔸Понедельник', callback_data="dz_1"), types.InlineKeyboardButton(text='🔸Вторник', callback_data="dz_2"))
    days.add(types.InlineKeyboardButton(text='🔸Среда', callback_data="dz_3"), types.InlineKeyboardButton(text='🔸Четверг', callback_data="dz_4"))
    days.add(types.InlineKeyboardButton(text='🔸Пятница', callback_data="dz_5"), types.InlineKeyboardButton(text='🔸Суббота', callback_data="dz_6"))
    days.add(types.InlineKeyboardButton(text='Следующая неделя >>', callback_data='dznext'))
    days.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='menu'))
    await bot.send_photo(call.from_user.id, 'https://imgur.com/c4WZxTm.jpg', '*🔀Выберите* день *текущей* недели:', parse_mode='markdown', reply_markup=days)

@bot.callback_query_handler(lambda call: call.data[0:3]=="dz_" and len(call.data)==4)
async def dz_n(call):
    try:
        n = int(call.data[3])
        rowrt = cursor.execute("SELECT * FROM credits WHERE user_id=?", (call.from_user.id,)).fetchone()
        if rowrt is None:
            await bot.answer_callback_query(call.id, '❎Вы не авторизованы', show_alert=True)
            return

        async with ElschoolClient(rowrt[1], rowrt[2]) as client:
            status0, t0 = await client.auth()
            if not status0:
                await bot.send_message(call.from_user.id, '❎Не удалось войти в аккаунт Elschool. Возможно, стоит обновить cookies?')
                return

            status1, t1 = await client.get_url('https://elschool.ru/users/diaries')
            if not status1:
                await bot.send_message(call.from_user.id, '❎Не удалось загрузить страницу дневника. Возможно, стоит обновить cookies?')
                return

            s = t1
            #print(s)
            s1 = s.split('<tbody>')[1:]
            tbody = s1[n-1]
            s = ''
            s2 = tbody.split('<tr class="diary__lesson">')[1:]
            s22 = s2.copy()
            date = s22[0].split('<td class="diary__dayweek')[1].split('<p>')[1].split('</p>')[0].replace('&nbsp;&nbsp;&nbsp;', ' ')
            #print(len(s2))
            j = 0
            for lesson in s2:
                j+=1
                name = lesson.split('<div class="flex-grow-1">')[1].split('</div>')[0]
                work = lesson.split('<div class="diary__homework-text">')[1].split('</div>')[0].replace("&#167;", "§").replace("&quot;", '"').replace("<", "&lt;").replace(">", "&gt;").replace("&", "&amp;")
                flair = "✍"
                if not work:
                    work='<em>ДЗ отсутствует</em>'
                    flair = ""
                s+=f"<ins>{j}.) {name.split('.')[1][1:]}</ins>: {flair}{work};\n"
            rm = types.InlineKeyboardMarkup()
            rm.add(types.InlineKeyboardButton(text='🔙Назад', callback_data='dz'))
            await bot.send_photo(call.from_user.id, 'https://imgur.com/m0B8XWn.jpg', f'<strong>📚Домашнее задание</strong> на {date}\n\n{s}', parse_mode='html', reply_markup=rm)
    except:
        #print(traceback.format_exc())
        pass


@bot.callback_query_handler(lambda call: call.data == 'dznext')
async def dz(call):
    try:
        await bot.delete_message(call.message.chat.id, call.message.message_id)
    except:
        pass
    days = types.InlineKeyboardMarkup()
    days.add(types.InlineKeyboardButton(text='🔸Понедельник', callback_data="dznext_1"), types.InlineKeyboardButton(text='🔸Вторник', callback_data="dznext_2"))
    days.add(types.InlineKeyboardButton(text='🔸Среда', callback_data="dznext_3"), types.InlineKeyboardButton(text='🔸Четверг', callback_data="dznext_4"))
    days.add(types.InlineKeyboardButton(text='🔸Пятница', callback_data="dznext_5"), types.InlineKeyboardButton(text='🔸Суббота', callback_data="dznext_6"))
    days.add(types.InlineKeyboardButton(text='<< Предыдущая неделя', callback_data='dz'))
    days.add(types.InlineKeyboardButton(text='🔙 Назад', callback_data='menu'))
    await bot.send_photo(call.from_user.id, 'https://imgur.com/c4WZxTm.jpg', '*🔀Выберите* день *следующей* недели:', parse_mode='markdown', reply_markup=days)


@bot.callback_query_handler(lambda call: call.data[0:7] == 'dznext_' and len(call.data) == 8)
async def dznext_n(call):
    try:
        n = int(call.data[7])
        rowrt = cursor.execute("SELECT * FROM credits WHERE user_id=?", (call.from_user.id,)).fetchone()
        if rowrt is None:
            await bot.answer_callback_query(call.id, '❎ Вы не авторизованы', show_alert=True)
            return

        async with ElschoolClient(rowrt[1], rowrt[2]) as client:
            status0, t0 = await client.auth()
            if not status0:
                await bot.send_message(call.from_user.id, '❎ Не удалось войти в аккаунт Elschool. Возможно, стоит обновить cookies?')
                return

            status1, t1 = await client.get_url('https://elschool.ru/users/diaries')
            if not status1:
                await bot.send_message(call.from_user.id, '❎ Не удалось загрузить страницу дневника. Возможно, стоит обновить cookies?')
                return

            s = t1
            ur = s.split('<div class="navigation__nextweek">')[1].split('<a href="')[1].split('"')[0].replace("&amp;", "&")
            status2, t2 = await client.get_url(f'https://elschool.ru{ur}')
            if not status2:
                await bot.send_message(call.from_user.id, '❎ Не удалось загрузить страницу дневника на следующую неделю. Возможно, стоит обновить cookies?')
                return

            s = t2
            #print(s)
            s1 = s.split('<tbody>')[1:]
            tbody = s1[n-1]
            s = ''
            s2 = tbody.split('<tr class="diary__lesson">')[1:]
            s22 = s2.copy()
            date = s22[0].split('<td class="diary__dayweek')[1].split('<p>')[1].split('</p>')[0].replace('&nbsp;&nbsp;&nbsp;', ' ')
            #print(len(s2))
            j = 0
            for lesson in s2:
                j+=1
                name = lesson.split('<div class="flex-grow-1">')[1].split('</div>')[0]
                work = lesson.split('<div class="diary__homework-text">')[1].split('</div>')[0].replace("&#167;", "§").replace("&quot;", '"').replace("<", "&lt;").replace(">", "&gt;").replace("&", "&amp;")
                flair = "✍"
                if not work:
                    work='<em>ДЗ отсутствует</em>'
                    flair = ""
                s+=f"<ins>{j}.) {name.split('.')[1][1:]}</ins>: {flair}{work};\n"
            rm = types.InlineKeyboardMarkup()
            rm.add(types.InlineKeyboardButton(text='🔙Назад', callback_data='dznext'))
            await bot.send_photo(call.from_user.id, 'https://imgur.com/m0B8XWn.jpg', f'<strong>📚Домашнее задание</strong> на {date}\n\n{s}', parse_mode='html', reply_markup=rm)
    except:
        #print(traceback.format_exc())
        pass


# В САМОМ КОНЦЕ, ЧТОБЫ НЕ ПЕРЕХВАТЫВАТЬ КОМАНДЫ
@bot.message_handler(content_types=['text'])
async def text_message_func(message):
    r = cursor.execute("SELECT * FROM states WHERE user_id=?", (message.from_user.id,)).fetchone()
    if r is None:
        return

    if r[1] == 'login-and-password':
        sp = list(message.text.split('\n'))
        if len(sp) == 2 and not '<' in sp[0] and not '>' in sp[0] and not '<' in sp[1] and not '>' in sp[1] and not '&' in sp[1] and not '&' in sp[0]:
            async with ElschoolClient(sp[0], sp[1]) as client:
                status, t = await client.auth()
                if status:
                    cursor.execute("DELETE FROM credits WHERE user_id=?", (message.from_user.id,))
                    connect.commit()

                    cursor.execute("INSERT INTO credits VALUES(?,?,?);", [message.from_user.id, sp[0], sp[1]])
                    connect.commit()

                    cursor.execute("DELETE FROM states WHERE user_id=?", (message.from_user.id,))
                    connect.commit()

                    #print(t)
                    name = t.split('"personal-data__name">')[1].split('<')[0].strip().lstrip()

                    zx = types.InlineKeyboardMarkup()
                    zx.add(types.InlineKeyboardButton(text='🔙 В меню', callback_data='menu'))
                    await bot.send_message(message.from_user.id, f'✅ Удалось войти в аккаунт Elschool\n\nПользователь: {name}\n\nИспользуйте команду /get_marks для загрузки оценок', reply_markup=zx)
                else:
                    zx = types.InlineKeyboardMarkup()
                    zx.add(types.InlineKeyboardButton(text='🔙 В меню', callback_data='menu'))
                    await bot.send_message(message.from_user.id, '❎ Не удалось войти в аккаунт Elschool. Попробуйте ещё раз👇', reply_markup=zx)

        else:
            await bot.reply_to(message, 'Некорректный ввод. Попробуйте ещё раз. По образцу.\n\n_aaaaa_\n_bbbbb_', parse_mode='markdown')


async def parsing_loop():
    global USERS_IN_CYCLE, AVG_PER_USER
    while True:
        try:
            # Получаем всех пользователей
            users = cursor.execute("SELECT user_id FROM users_posting").fetchall()

            USERS_IN_CYCLE = len(users)
            # print(USERS_IN_CYCLE)
            total = 0

            for i, (user_id,) in enumerate(users):
                # print(user_id)
                user_start_time = time.time()
                try:
                    f = await process_marks(user_id)
                    if f is False:
                        cursor.execute("DELETE FROM users_posting WHERE user_id=?", (user_id,))
                        connect.commit()

                        await bot.send_message(user_id, '❎ Вы были исключены из цикла автоматической проверки оценок')

                    user_time = round(time.time() - user_start_time, 2)
                    total += user_time
                    AVG_PER_USER = total / (i + 1)

                    # Небольшая пауза между пользователями
                    if i % 10 == 0:  # Каждого 10-го пользователя ждем чуть дольше
                        await asyncio.sleep(5)
                    else:
                        await asyncio.sleep(2)

                except Exception as e:
                    print(f"❌ Ошибка у пользователя {user_id}: {e}")
                    continue

            await asyncio.sleep(60)

        except Exception as e:
            print(f"🔥 Критическая ошибка в цикле: {e}")
            await asyncio.sleep(60)  # Пауза при ошибке


async def main():
    # Запускаем парсинг в фоне
    asyncio.create_task(parsing_loop())

    # Запускаем бота
    await bot.polling(none_stop=True, interval=0)


if __name__ == "__main__":
    asyncio.run(main())

# asyncio.run(bot.polling(none_stop=True, interval=0))
